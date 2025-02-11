from flask import Flask, render_template, request, url_for,redirect, flash, jsonify,session,send_from_directory, send_file
from database import db_sql_server
from flask_login import LoginManager,login_user, logout_user, login_required,current_user
from flask_wtf.csrf import CSRFProtect
import uuid
from datetime import datetime, timedelta, date, time
import pytz


# Zona horaria de México (UTC-6)
mexico_tz = pytz.timezone('America/Mexico_City')

# Obtener la hora actual en UTC
utc_now = datetime.now(pytz.utc)

# Convertir a la zona horaria de México
mexico_now = utc_now.astimezone(mexico_tz)

from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import A4
from werkzeug.exceptions import BadRequestKeyError
import pdfkit
from flask import make_response
import base64
import os
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side

# Models
from models.ModelEmpresas import ModelEmpresas
from models.ModelClientes import Modelclientes
from models.ModelCategoria import ModelCategoria
from models.ModelPuesto import ModelPuesto
from models.ModelBanco import ModelBanco
from models.ModelRegistroPatronal import ModelRegistroPatronal
from models.ModelCuentaEmpresa import ModelCuentasEmpresas
from models.ModelCuentaCliente import ModelCuentasClientes
from models.ModelContactoCliente import ModelContactosClientes
from models.ModelCFDI import ModelCFDI
from models.ModelEspecialidades import ModelEspecialidades
from models.ModelProveedores import ModelProveedores
from models.ModelEspecialidadesProveerdor import ModelEspecialidadesProveedor
from models.ModelContactoProveedor import ModelContactosProveedor
from models.ModelCuentaProveedor import ModelCuentasProveedor
from models.ModelProyectoObra import ModelProyectoObra
from models.ModelContratistas import Modelcontratistas
from models.ModelFamilias import ModelFamilias
from models.ModelMaterialesFamilia import ModelMaterialesFamilia
from models.ModelMateriales import ModelMateriales
from models.ModelEmpleado import ModelEmpleado
from models.ModelMonedero import ModelMonedero
from models.ModelAsignaciones import ModelAsignaciones
from models.ModelUser import ModelUser
from models.ModelRequisiciones import ModelRequisiciones
from models.ModelPresupuesto import ModelPresupuesto
from models.ModelAsistencias import AsistenciaModel
from models.ModelRequisiciones import ModelRequisiciones
from models.ModelIncidencias import IncidenciaModel
from models.ModelNomina import ModelNomina
from models.ModelOrden import MyOrdendeCompra
from models.ModelPDF import ModelPDF
from models.ModelCompras import ModelCompras

# Entities
from models.entities.Empresas import Empresas
from models.entities.Nomina import Nomina
from models.entities.Asignacion import Asignacion
from models.entities.Clientes import Clientes
from models.entities.Categoria import Categoria
from models.entities.Puestos import Puesto
from models.entities.Banco import Banco
from models.entities.Incidencias import Incidencia
from models.entities.RegistroPatronal import RegistroPatronal
from models.entities.CuentasEmpresas import CuentasEmpresas
from models.entities.CuentasClientes import CuentasClientes
from models.entities.ContactoClientes import ContactosClientes
from models.entities.Monedero import Monedero
from models.entities.UsoCFDI import UsoCFDI
from models.entities.Especialidades import Especialidades
from models.entities.Proveedores import Proveedores
from models.entities.EspecialidadesProveedor import EspecialidadesProveedor
from models.entities.ContactoProveedores import ContactosProveedores
from models.entities.CuentasProveedor import CuentasProveedor
from models.entities.ProyectoObra import ProyectoObra
from models.entities.Contratistas import Contratistas
from models.entities.Familias import Familias
from models.entities.MaterialesFamilia import MaterialesFamilia
from models.entities.Empleados import Empleados
from models.entities.Usuarios import User
from models.entities.Presupuestos import Presupuesto,DetallePresupuesto,PresupuestoBauart,DetalleBauart
from models.entities.Materiales import Materiales

app = Flask(__name__)



# CONEXIÓN BASE DE DATOS
db = db_sql_server.get_connection()

app.config['SECRET_KEY'] = '62AC6398Jf8VxZCR'
csrf = CSRFProtect()
csrf.init_app(app)


login_manager = LoginManager()
login_manager.init_app(app) 

@login_manager.user_loader
def load_user(id):
    login_user =  ModelUser.get_by_id(db,id)

    return login_user


@app.route('/logout')
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/')
def index():
    return redirect(url_for('login'))


@app.route('/favicon.ico')
def favicon():
    return send_from_directory('static', 'Buart.ico', mimetype='image/vnd.microsoft.icon')
        

@app.route('/login', methods=['GET', 'POST'])
def login():
    try:
        empresas = ModelEmpresas.get_all_empresas(db)
    except Exception as e:
        raise e

    if request.method == 'POST':
        usuario = request.form['USUARIO']
        password = request.form['CONTRASENA']
        empresa = request.form['EMPRESA']

        user = User(0, empresa, usuario, "", password, "", False)
        logged_user = ModelUser.login(db, user)
        
        if logged_user:
            if logged_user.password:
                # Login del usuario en la sesión
                login_user(logged_user)
                return redirect(url_for('home'))
            
            else:
                flash("Contraseña incorrectos")
                return render_template('login.html', empresas=empresas)
                
        else:
            flash("Usuario incorrectos")
            return render_template('login.html', empresas=empresas)
    else:
        return render_template('login.html', empresas=empresas)
    
@app.route('/Home')
@login_required
def home():
    if current_user.is_authenticated:
        print(f"ESTA ES EL USAURIO {current_user.id}")
        return render_template('home.html')
    
    else:
        flash("Ocurrió un error al intentar iniciar sesión.")
        return redirect(url_for('login'))
        
        
@app.route('/Periodos')
def periodos():
    return render_template('periodos.html')


"""
INICIA EMPRESAS
"""

#ALTA EMPRESA
@app.route('/Altaempresas',methods=['GET', 'POST'])
@login_required
def altaempresas():
    if request.method == 'POST':
        
        # Verifica si el token en la sesión coincide con el del formulario
        token = request.form.get('token')
        if token and session.get('token') == token:
            # Elimina el token de la sesión para evitar reenvíos duplicados
            session.pop('token', None)
        
            new_empresa = Empresas(
                id=request.form['ID_EMPRESA'],
                repse=request.form['REPSE'],
                razon_social=request.form['RAZON_SOCIAL'],
                rfc=request.form['RFC'],
                nombre_representante1=request.form['NOMBRE_REPRESENTANTE1'],
                apellido_representante1=request.form['APELLIDO_REPRESENTANTE1'],
                telefono_representante1=request.form['TELEFONO_REPRESENTANTE1'],
                correo_representante1=request.form['CORREO_REPRESENTANTE1'],
                calle=request.form['CALLE'],
                no_exterior= request.form.get('NO_EXTERIOR', ""),
                no_interior= request.form.get('NO_INTERIOR', ""),
                cp=request.form['CP'],
                estado=request.form['ESTADO'],
                municipio=request.form['MUNICIPIO'],
                pais=request.form['PAIS'],
                regimen_fiscal_id=request.form['REGIMEN_FISCAL'],
                nombre_representante2=request.form.get('NOMBRE_REPRESENTANTE2', ""),
                apellido_representante2=request.form.get('APELLIDO_REPRESENTANTE2', ""),
                telefono_representante2=request.form.get('TELEFONO_REPRESENTANTE2', ""),
                correo_representante2=request.form.get('CORREO_REPRESENTANTE2', ""),
                nombre_apoderado=request.form.get('NOMBRE_APODERADO', ""),
                apellido_apoderado=request.form.get('APELLIDO_APODERADO', ""),
                telefono_apoderado=request.form.get('TELEFONO_APODERADO', ""),
                correo_apoderado=request.form.get('CORREO_APODERADO', "")
               
            )

        
            ModelEmpresas.crearEmpresa(db, new_empresa)
            
            estados = request.form.getlist('estado[]')
            registros_patronales = request.form.getlist('noRegistro[]')

            
            # Verifica que ambos tengan el mismo tamaño
            if len(estados) == len(registros_patronales):
                for i in range(len(estados)):
                    nuevo_regis = RegistroPatronal(id_registro=0,
                                                id_empresa=request.form['ID_EMPRESA'],
                                                numero_registro_patronal=registros_patronales[i],
                                                estado=estados[i]
                    )
                    ModelRegistroPatronal.newRegistroPatronal(db,nuevo_regis)


            bancos = request.form.getlist('BANCO[]')
            numero_cuenta = request.form.getlist('NO_CUENTA[]')
            clabes = request.form.getlist('CLABE[]')

            for i in range(len(bancos)):
                nueva_cuenta = CuentasEmpresas(
                    id=0,
                    id_empresa=request.form['ID_EMPRESA'],
                    id_banco=bancos[i],  # Accede directamente por índice
                    numero_cuenta=numero_cuenta[i],
                    clabe=clabes[i]
                )
                ModelCuentasEmpresas.new_cuenta_empresa(db, nueva_cuenta)

            return redirect(url_for('empresas'))

        else:
            
            return redirect(url_for('empresas'))
        
    else:
        
        # Genera un token único y lo almacena en la sesión
        session['token'] = str(uuid.uuid4())
        
        ultima_empresa = ModelEmpresas.get_empresa_last_by_id(db)
        if ultima_empresa == None:
            ultima_empresa = 1

        else:
            ultima_empresa = ultima_empresa + 1
            
        return render_template('altaEmpresas.html',ultima_empresa = ultima_empresa,token=session['token'])

#VALIDAR RFC DE EMPRESAS
@app.route('/api/validar_rfc_empresas/', methods=['GET'])
@login_required
def validar_rfc_empresas():
    try:
        rfc = request.args.get('rfc', '')

        empresas = ModelEmpresas.get_empresas_not_block(db)

        existe_rfc = False

        for e in empresas:
            if e.rfc == rfc:
                existe_rfc = True
              
            
        return jsonify(existe_rfc), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
#EMPRESAS 
@app.route('/Empresas', methods=['GET', 'POST'])
@login_required
def empresas():
    if request.method == 'POST':
         # Verifica si el token en la sesión coincide con el del formulario
        token = request.form.get('token')
        if token and session.get('token') == token:
            # Elimina el token de la sesión para evitar reenvíos duplicados
            session.pop('token', None)
            
        repse = request.form.get('REPSE', '')
        razon_social = request.form.get('RAZON_SOCIAL', '')
        rfc = request.form.get('RFC', '')
        cp = request.form.get('CP', '')
        estado = request.form.get('ESTATUS', '')

        empresas = ModelEmpresas.filter_empresas(db, repse, razon_social, rfc, cp,estado)
        return render_template('empresas.html', empresas=empresas)
    
    
    empresa = ModelEmpresas.get_all_empresas(db)
    # Genera un token único y lo almacena en la sesión
    session['token'] = str(uuid.uuid4())

    return render_template('empresas.html', empresas=empresa,token=session['token'])

#EDIT EMPRESA
@app.route('/edit_empresa/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_empresa(id):

    if request.method == 'POST':
        # Verifica si el token en la sesión coincide con el del formulario
        token = request.form.get('token')
        if token and session.get('token') == token:
            # Elimina el token de la sesión para evitar reenvíos duplicados
            session.pop('token', None)
            
            empresa = Empresas(
                id=id,
                repse=request.form['REPSE'],
                razon_social=request.form['RAZON_SOCIAL'],
                rfc=request.form['RFC'],
                nombre_representante1=request.form['NOMBRE_REPRESENTANTE1'],
                apellido_representante1=request.form['APELLIDO_REPRESENTANTE1'],
                telefono_representante1=request.form['TELEFONO_REPRESENTANTE1'],
                correo_representante1=request.form['CORREO_REPRESENTANTE1'],
                calle=request.form['CALLE'],
                no_exterior= request.form.get('NO_EXTERIOR', ""),
                no_interior= request.form.get('NO_INTERIOR', ""),
                cp=request.form['CP'],
                estado=request.form['ESTADO'],
                municipio=request.form['MUNICIPIO'],
                pais=request.form['PAIS'],
                regimen_fiscal_id=request.form['REGIMEN_FISCAL'],
                nombre_representante2=request.form.get('NOMBRE_REPRESENTANTE2', ""),
                apellido_representante2=request.form.get('APELLIDO_REPRESENTANTE2', ""),
                telefono_representante2=request.form.get('TELEFONO_REPRESENTANTE2', ""),
                correo_representante2=request.form.get('CORREO_REPRESENTANTE2', ""),
                nombre_apoderado=request.form.get('NOMBRE_APODERADO', ""),
                apellido_apoderado=request.form.get('APELLIDO_APODERADO', ""),
                telefono_apoderado=request.form.get('TELEFONO_APODERADO', ""),
                correo_apoderado=request.form.get('CORREO_APODERADO', "")
               
            )
            
            ModelEmpresas.update_empresa(db, empresa)
            ModelRegistroPatronal.delete_registro_patronales(db, empresa.id)

            estados = request.form.getlist('estado[]')
            registros_patronales = request.form.getlist('noRegistro[]')

            
            # Verifica que ambos tengan el mismo tamaño
            if len(estados) == len(registros_patronales):
                for i in range(len(estados)):
                    nuevo_regis = RegistroPatronal(id_registro=0,
                                                id_empresa=request.form['ID_EMPRESA'],
                                                numero_registro_patronal=registros_patronales[i],
                                                estado=estados[i]
                    )
                    ModelRegistroPatronal.newRegistroPatronal(db,nuevo_regis)

            ModelCuentasEmpresas.delete_cuentas_empresa(db,id)

            bancos = request.form.getlist('BANCO[]')
            numero_cuenta = request.form.getlist('NO_CUENTA[]')
            clabes = request.form.getlist('CLABE[]')

            for i in range(len(bancos)):
                nueva_cuenta = CuentasEmpresas(
                    id=0,
                    id_empresa=request.form['ID_EMPRESA'],
                    id_banco=bancos[i],  # Accede directamente por índice
                    numero_cuenta=numero_cuenta[i],
                    clabe=clabes[i]
                )
                ModelCuentasEmpresas.new_cuenta_empresa(db, nueva_cuenta)

                return redirect(url_for('empresas'))
            
        else:
            return redirect(url_for('empresas'))
            
    
    else:
        # Genera un token único y lo almacena en la sesión
        session['token'] = str(uuid.uuid4())

        empresa = ModelEmpresas.get_empresa_by_id(db, id)
        registrosPatronales = ModelRegistroPatronal.get_RegistroPatronal_empresa(db,empresa.id)
        cuentas = ModelCuentasEmpresas.get_cuenta_empresa_by_id(db,id)
        bancos = ModelBanco.get_all_bancos(db)
    
        return render_template('edit_empresa.html', empresa=empresa, registrosPatronales=registrosPatronales,cuentas=cuentas, bancos=bancos,token=session['token'])

#BLOQUEAR EMPRESA
@app.route('/block_empresa/<int:id>', methods=['POST'])
@login_required
def block_empresa(id):
    if request.method == 'POST':
        try:
       
            ModelEmpresas.change_status(db, id, True)
            return redirect(url_for('empresas'))
        except Exception as e:
            return jsonify({'error': str(e)}), 500

#DESBLOQUEA EMPRESA
@app.route('/unblock_empresa/<int:id>', methods=['POST'])
@login_required
def unblock_empresa(id):
    if request.method == 'POST':
        try:
          
            ModelEmpresas.change_status(db, id, False)
            return redirect(url_for('empresas'))
        except Exception as e:
            return jsonify({'error': str(e)}), 500
            
#BLOCK REGISTRO PATRONAL
@app.route('/block_registro_patronal/<int:id>', methods=['POST'])
def block_registro_patronal(id):
    if request.method == 'POST':
        try:
         
            ModelRegistroPatronal.change_status(db, id, True)
            return redirect(url_for('registosPatronales'))
        except Exception as e:
            return jsonify({'error': str(e)}), 500
            
#UNBLOCK REGISTRO PATRONAL
@app.route('/unblock_registro_patronal/<int:id>', methods=['POST'])
def unblock_registro_patronal(id):
    if request.method == 'POST':
        try:
            return jsonify({'error': str(e)}), 500
            ModelRegistroPatronal.change_status(db, id, False)
            return redirect(url_for('registosPatronales'))
        except Exception as e:
            return jsonify({'error': str(e)}), 500
            
#REGISTROS PATRONALES
@app.route('/RegistosPatronales', methods=['GET', 'POST'])
def registosPatronales():
    if request.method == 'POST':
        new_regis_patronal = RegistroPatronal(id_registro=0,
                                  registro_patronal=request.form['NO_REGISTRO_PATRONAL'],
                                  empresa=request.form['EMPRESA'],
                                  estado=request.form['ESTADO_REGISTRO']
                                  )
        
      
        ModelRegistroPatronal.newRegistroPatronal(db,new_regis_patronal)
        return redirect(url_for('registosPatronales'))
        
        
    empresas = ModelEmpresas.get_all_empresas(db)
    regis_patronales = ModelRegistroPatronal.get_all_registroPatronal(db)

    
    return render_template('registrosPatronales.html', regis_patronales = regis_patronales, empresas=empresas)

# EDIT REGISTRO PATRONAL
@app.route('/edit_registro_patronal/<int:id>', methods=['GET', 'POST'])
def edit_registro_patronal(id):
    if request.method == 'POST':
        regis_patronal = RegistroPatronal(
            id_registro = id,
            registro_patronal=request.form['NO_REGISTRO_PATRONAL'],
            empresa=request.form['ID_EMPRESA'],
            estado= request.form['ESTADO_REPRESENTANTE']   
        )
        
        ModelRegistroPatronal.update_registro_patronal(db, regis_patronal)
        return redirect(url_for('registosPatronales'))


"""
FIN EMPRESAS 
"""

"""
INICIA CLIENTES
"""

#ALTA CLIENTE 
@app.route('/Altacliente',methods=['GET', 'POST'])
def altaclientes():

    if request.method == 'POST':
        # Verifica si el token en la sesión coincide con el del formulario
        token = request.form.get('token')
        if token and session.get('token') == token:
            # Elimina el token de la sesión para evitar reenvíos duplicados
            session.pop('token', None)
            
            cliente = Clientes(id = request.form['ID_CLIENTE'],
                            id_empresa=request.form['EMPRESA'],
                            razon_social=request.form['RAZON_SOCIAL'],
                            rfc=request.form['RFC'],
                            cp=request.form['CP'],
                            estado=request.form['MUNICIPIO'],
                            ciudad=request.form['PAIS'],
                            municipio=request.form['CALLE'],
                            pais=request.form['PAIS'],
                            calle=request.form['CALLE'],
                            no_exterior=request.form.get('NO_EXTERIOR', ""),
                            regimen_fiscal_id=request.form['REGIMEN_FISCAL'],
                            uso_cfdi_id=request.form['USO_CFDI'],
                            condicion_pago=request.form['CONDICION_PAGO'],
                            forma_pago_id= request.form['FORMA_PAGO']             
            )
                            
            Modelclientes.crearCliente(db, cliente)

            nombres_contactos = request.form.getlist('NOMBRE_CONTACTO[]')
            apellidos_coctactos = request.form.getlist('APELLIDO_CONTACTO[]')
            telefonos_contacto = request.form.getlist('TELEFONO_CONTACTO[]')
            correos_contacto = request.form.getlist('CORREO_CONTACTO[]')
            puestos_contacto = request.form.getlist('PUESTO_CONTACTO[]')

            
            for i in range(len(nombres_contactos)):
                nuevo_contacto = ContactosClientes(
                    id = 0,
                    id_cliente= cliente.id,
                    nombre= nombres_contactos[i],
                    apellido= apellidos_coctactos[i],
                    puesto=puestos_contacto[i],
                    telefono=telefonos_contacto[i],
                    correo=correos_contacto[i],
                    usuario=current_user.id
                )

                ModelContactosClientes.new_contacto_cliente(db,nuevo_contacto)
            
            bancos = request.form.getlist('ID_BANCO[]')
            numeros_cuentas = request.form.getlist('NUM_CUENTA[]')
            clabes = request.form.getlist('CLABE[]')

            for b in range(len(bancos)):
                nueva_cuenta = CuentasClientes(
                    id=0,
                    id_cliente=cliente.id,
                    id_banco=bancos[b],
                    numero_cuenta=numeros_cuentas[b],
                    clabe=clabes[b],
                    usuario=current_user.id,
                )

                ModelCuentasClientes.new_cuenta_cliente(db,nueva_cuenta)
        
            return redirect(url_for('clientes'))
        else:
            
            return redirect(url_for('clientes'))
    
    else:
        # Genera un token único y lo almacena en la sesión
        session['token'] = str(uuid.uuid4())
        
        ultimo_cliente = Modelclientes.get_cliente_last_by_id(db)
        if ultimo_cliente is None:
            ultimo_cliente = 1
           
        else:
            ultimo_cliente = ultimo_cliente + 1
        
        empresas = ModelEmpresas.get_empresas_not_block(db)
        return render_template('altaClientes.html',empresas=empresas,ultimo_cliente=ultimo_cliente,token=session['token'])

#VALIDAR RFC DE CLIENTES
@app.route('/api/validar_rfc_clientes/', methods=['GET'])
def validar_rfc_clientes():
    
    try:
        rfc = request.args.get('rfc', '')

        cliente = Modelclientes.get_clientes_not_block(db)

        existe_rfc = False

        for c in cliente:
            if c.rfc == rfc:
                existe_rfc = True
              
            
        return jsonify(existe_rfc), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500

# EDIT CLIENTE
@app.route('/edit_cliente/', methods=['GET', 'POST'])
def edit_cliente():
    id = request.args.get('id')
    if request.method == 'POST':
         # Verifica si el token en la sesión coincide con el del formulario
        token = request.form.get('token')
        if token and session.get('token') == token:
            # Elimina el token de la sesión para evitar reenvíos duplicados
            session.pop('token', None)
            
            cliente = Clientes(id = id,
                            id_empresa=request.form['EMPRESA'],
                            razon_social=request.form['RAZON_SOCIAL'],
                            rfc=request.form['RFC'],
                            cp=request.form['CP'],
                            estado=request.form['MUNICIPIO'],
                            ciudad=request.form['PAIS'],
                            municipio=request.form['CALLE'],
                            pais=request.form['PAIS'],
                            calle=request.form['CALLE'],
                            no_exterior=request.form.get('NO_EXTERIOR', ""),
                            regimen_fiscal_id=request.form['REGIMEN_FISCAL'],
                            uso_cfdi_id=request.form['USO_CFDI'],
                            condicion_pago=request.form['CONDICION_PAGO'],
                            forma_pago_id= request.form['FORMA_PAGO']             
            )
            
            Modelclientes.update_cliente(db, cliente)
            
            nombres_contactos = request.form.getlist('NOMBRE_CONTACTO[]')
            apellidos_coctactos = request.form.getlist('APELLIDO_CONTACTO[]')
            telefonos_contacto = request.form.getlist('TELEFONO_CONTACTO[]')
            correos_contacto = request.form.getlist('CORREO_CONTACTO[]')
            puestos_contacto = request.form.getlist('PUESTO_CONTACTO[]')

            ModelContactosClientes.delete_contactos(db,id)
            
            for i in range(len(nombres_contactos)):
                nuevo_contacto = ContactosClientes(
                    id = 0,
                    id_cliente= cliente.id,
                    nombre= nombres_contactos[i],
                    apellido= apellidos_coctactos[i],
                    puesto=puestos_contacto[i],
                    telefono=telefonos_contacto[i],
                    correo=correos_contacto[i],
                    usuario=1
                )
               
                ModelContactosClientes.new_contacto_cliente(db,nuevo_contacto)
            
            bancos = request.form.getlist('ID_BANCO[]')
            numeros_cuentas = request.form.getlist('NUM_CUENTA[]')
            clabes = request.form.getlist('CLABE[]')

            for b in range(len(bancos)):
                nueva_cuenta = CuentasClientes(
                    id=0,
                    id_cliente=cliente.id,
                    id_banco=bancos[b],
                    numero_cuenta=numeros_cuentas[b],
                    clabe=clabes[b],
                    usuario=1,
                )
                ModelCuentasClientes.delete_cuentas(db,id)
                ModelCuentasClientes.new_cuenta_cliente(db,nueva_cuenta)
        
            return redirect(url_for('clientes'))
        
        else:            
            return redirect(url_for('clientes'))
            
        
    else:    
        # Genera un token único y lo almacena en la sesión
        session['token'] = str(uuid.uuid4())
        cliente = Modelclientes.get_cliente_by_id(db, id)
        empresa_selecionada = ModelEmpresas.get_empresa_by_id(db, cliente.id_empresa)
        empresas = ModelEmpresas.get_all_empresas(db)
        contactos = ModelContactosClientes.get_contactos_by_cliente(db, id)
        uso_cfdi = ModelCFDI.get_all_usoCFDI(db)
        cuentas = ModelCuentasClientes.get_all_cuentas_cliente(db,id)
        bancos = ModelBanco.get_all_bancos(db)
        return render_template('edit_cliente.html', cliente=cliente, empresas = empresas,contactos=contactos, empresa_selecionada=empresa_selecionada,uso_cfdi=uso_cfdi,cuentas=cuentas,bancos=bancos,token=session['token'])


#CLIENTES
@app.route('/Clientes',  methods=['GET', 'POST'])
def clientes():
    cliente = Modelclientes.get_all_clientes(db)
    return render_template('clientes.html', clientes=cliente)

#BLOQUEAR CLIENTE
@app.route('/block_cliente/<int:id>', methods=['POST'])
def block_cliente(id):
    if request.method == 'POST':
        try:
            Modelclientes.change_status(db, id, True)
            return redirect(url_for('clientes'))
        except Exception as e:
           return jsonify({'error': str(e)}), 500

#DESBLOQUEAR CLIENTE
@app.route('/unblock_cliente/<int:id>', methods=['POST'])
def unblock_cliente(id):
    if request.method == 'POST':
        try:
            Modelclientes.change_status(db, id, False)
            return redirect(url_for('clientes'))
        except Exception as e:
            jsonify({'error': str(e)}), 500




"""
FIN CLIENTES
"""


"""
INICIO PROYECTO 
"""

## ALTA PROYECTO

@app.route('/AltaProyectoObra', methods=['GET', 'POST'])
def alta_proyecto():
    if request.method == 'POST':
        
        # Verifica si el token en la sesión coincide con el del formulario
        token = request.form.get('token')
        if token and session.get('token') == token:
            # Elimina el token de la sesión para evitar reenvíos duplicados
            session.pop('token', None)
            
            nuevo_proyecto = ProyectoObra(
                        id=0,
                        id_cliente=request.form['cliente_id'],
                        id_empresa=current_user.id_empresa,
                        fecha_inicio=request.form['fecha_inicio'],
                        fecha_contrato="0000-00-00",
                        fecha_fin=request.form['fecha_fin'],
                        nombre_proyecto=request.form['nombre_proyecto'],
                        centro_comercial=request.form.get('centro_comercial', ''),
                        pais='MÉXICO',
                        estado=request.form['ESTADO'],
                        municipio=request.form['MUNICIPIO'],
                        colonia=request.form.get('COLONIA', ''),
                        calle=request.form['CALLE'],
                        numero_exterior=request.form['NUMERO_EXTERIOR'],
                        numero_interior=request.form.get('NUMERO_INTERIOR', ''),
                        director_proyecto=request.form['nombre_director'],
                        lider_proyecto=request.form['lider1_proyecto'],
                        gerente_proyecto=request.form.get('nombre_gerente', ''),  # Ajuste en caso de que no haya un campo gerente
                        lider1=request.form['lider1_proyecto'],
                        lider2=request.form.get('lider2_proyecto', ''),
                        cp = request.form['CP'],
                        tipo_id = request.form.get('probra', ''),
                        hora_entrada = request.form.get('HORA_ENTRADA', ''),
                        hora_salida = request.form.get('HORA_SALIDA', ''),
                        latitud=request.form.get('LATITUD', ''),
                        longitud=request.form.get('LONGITUD', ''),
                        direcion_obra=request.form.get('DIRECCION_OBRA', ''),
                        usuario_id=current_user.id
            )
            
            ModelProyectoObra.crear_ProyectoObra(db, nuevo_proyecto)
            return redirect(url_for('proyectos'))
        
        else:
              return redirect(url_for('proyectos'))
    
    else:  
        # Genera un token único y lo almacena en la sesión            
        session['token'] = str(uuid.uuid4())
        directores = ModelEmpleado.obtener_directores(db)
        lideres = ModelEmpleado.obtener_lideres(db)
        gerentes = ModelEmpleado.obtener_gerentes(db)
        
        return render_template('altaProyectoObra.html',token=session['token'],directores=directores,lideres=lideres,gerentes=gerentes)


#PROYECTO
@app.route('/Proyectos',  methods=['GET', 'POST'])
def proyectos():

    if request.method == 'POST':
         
        # Verifica si el token en la sesión coincide con el del formulario
        token = request.form.get('token')
        if token and session.get('token') == token:
            # Elimina el token de la sesión para evitar reenvíos duplicados
            session.pop('token', None)
            # Obtener valores del formulario
            proyecto = request.form.get('PROYECTO', '')
            id_cliente = request.form.get('CLIENTE_ID', '')
            estado = request.form.get('ESTATUS', '')
            # Llamar al método filter_clientes con los parámetros obtenidos del formulario
            proyectos = ModelProyectoObra.filter_proyecto(db,proyecto, id_cliente, estado)
            return render_template('proyectoobra.html', proyectos=proyectos)

        else:
            return redirect(url_for('proyectos'))
    else:
        # Genera un token único y lo almacena en la sesión
        session['token'] = str(uuid.uuid4())
        proyectos = ModelProyectoObra.get_all_proyectos(db)
        return render_template('proyectoobra.html', proyectos=proyectos,token=session['token'])

#BLOQUEAR PROYECTO
@app.route('/block_proyecto/<int:id>', methods=['POST'])
def block_proyecto(id):
    if request.method == 'POST':
        try:
       
            ModelProyectoObra.change_status(db, id, True)
            return redirect(url_for('proyectos'))
        except Exception as e:
            return jsonify({'error': str(e)}), 500

#DESBLOQUEA PROYECTO
@app.route('/unblock_proyecto/<int:id>', methods=['POST'])
def unblock_proyecto(id):

    if request.method == 'POST':
        try:
          
            ModelProyectoObra.change_status(db, id, False)
            return redirect(url_for('proyectos'))
        except Exception as e:
            return jsonify({'error': str(e)}), 500



# EDITAR PROYECTO
@app.route('/edit_proyecto/<int:id>', methods=['GET', 'POST'])
def edit_proyecto(id):
    if request.method == 'POST':
        
        # Verifica si el token en la sesión coincide con el del formulario
        token = request.form.get('token')
        if token and session.get('token') == token:
            # Elimina el token de la sesión para evitar reenvíos duplicados
            session.pop('token', None)
            nuevo_proyecto = ProyectoObra(
                        id=id,
                        id_empresa=current_user.id_empresa,
                        id_cliente=request.form['cliente_id'],
                        fecha_inicio=request.form['fecha_inicio'],
                        fecha_contrato="2024-08-22",
                        fecha_fin=request.form['fecha_fin'],
                        nombre_proyecto=request.form['nombre_proyecto'],
                        centro_comercial=request.form.get('centro_comercial', ''),
                        pais='MÉXICO',
                        estado=request.form['ESTADO'],
                        municipio=request.form['MUNICIPIO'],
                        colonia=request.form.get('COLONIA', ''),
                        calle=request.form['CALLE'],
                        numero_exterior=request.form['NUMERO_INTERIOR'],
                        numero_interior=request.form.get('NUMERO_EXTERIOR', ''),
                        director_proyecto=request.form['nombre_director'],
                        lider_proyecto=request.form['lider1_proyecto'],
                        gerente_proyecto=request.form.get('nombre_gerente', ''),  # Ajuste en caso de que no haya un campo gerente
                        lider1=request.form['lider1_proyecto'],
                        lider2=request.form.get('lider2_proyecto', ''),
                        cp = request.form['CP'],
                        tipo_id = request.form.get('probra', ''),
                        hora_entrada = request.form.get('HORA_ENTRADA', ''),
                        hora_salida = request.form.get('HORA_SALIDA', ''),
                        latitud=request.form.get('LATITUD', ''),
                        longitud=request.form.get('LONGITUD', ''),
                        direcion_obra=request.form.get('DIRECCION_OBRA', ''),
                        usuario_id=current_user.id
            )
            
            ModelProyectoObra.update_proyecto(db, nuevo_proyecto)
            return redirect(url_for('proyectos'))
        
        else:
            return redirect(url_for('proyectos'))
    else:
        directores = ModelEmpleado.obtener_directores(db)
        lideres = ModelEmpleado.obtener_lideres(db)
        gerentes = ModelEmpleado.obtener_gerentes(db)
        # Genera un token único y lo almacena en la sesión
        session['token'] = str(uuid.uuid4())      
        proyecto = ModelProyectoObra.get_proyecto_by_id(db,id)   
        cleinte = Modelclientes.get_cliente_by_id(db,proyecto.id_cliente)
        return render_template('edit_proyecto.html', proyecto=proyecto,cliente=cleinte,token=session['token'],directores=directores,lideres=lideres,gerentes=gerentes)


    

"""
FIN PROYECTO 
"""
"""
INICIA EMPLEADOS
"""
#EMPLEADOS
@app.route('/Empleados',methods=['GET','POST'])
def empleados():
    try:
        if request.method == 'POST':
            # Verifica si el token en la sesión coincide con el del formulario
            token = request.form.get('token')
            if token and session.get('token') == token:
                # Elimina el token de la sesión para evitar reenvíos duplicados
                session.pop('token', None)
            
                nombre = request.form['NOMBRE']
                apellido = request.form['APELLIDO']
                tipo_empleado = request.form['TIPO_EMPLEADO']
                estatus = request.form['ESTATUS']

                empleados = ModelEmpleado.filter_empleados(db,nombre,apellido,tipo_empleado,estatus)

                for e in empleados:
                    puesto = ModelPuesto.get_puestos_by_id(db,e.puesto)
                    e.puesto = puesto.puesto 

                # Diccionario que mapea los valores de tipo_empleado a su descripción
                tipo_empleado_dict = {
                    1: "ADMINISTRATIVO",
                    2: "OBRA"  # Puedes agregar más tipos si es necesario
                }

                for e in empleados:
                    e.tipo_empleado = tipo_empleado_dict.get(e.tipo_empleado, "")
                        
                return render_template('empleados.html',empleados=empleados)
            
            else:
                return redirect(url_for('empleados'))
        
        else:
            # Genera un token único y lo almacena en la sesión
            session['token'] = str(uuid.uuid4())
            empleados = ModelEmpleado.get_all_empleados(db)   

            for e in empleados:
                puesto = ModelPuesto.get_puestos_by_id(db,e.puesto)
                e.puesto = puesto.puesto 

            # Diccionario que mapea los valores de tipo_empleado a su descripción
            tipo_empleado_dict = {
                1: "ADMINISTRATIVO",
                2: "OBRA"  # Puedes agregar más tipos si es necesario
            }

            for e in empleados:
                e.tipo_empleado = tipo_empleado_dict.get(e.tipo_empleado, "")
                    
            return render_template('empleados.html',empleados=empleados,token=session['token'])
        
    except Exception as ex:
        raise Exception(ex)

@app.route('/Altaempleados', methods=['GET', 'POST'])
def Altaempleados():
    try:
        print("Iniciando ruta /Altaempleados")

        # Obtener datos iniciales
        print("Obteniendo datos iniciales...")
        puesto = ModelPuesto.get_all_puestos_no_block(db)
        bancos = ModelBanco.get_all_bancos_no_block(db)
        empresas = ModelEmpresas.get_empresas_not_block(db)
        registosPatronales = ModelRegistroPatronal.get_all_registros_patronales(db)

        # Validar que los datos no estén vacíos
        if not puesto or not bancos or not empresas or not registosPatronales:
            return redirect(url_for('empleados'))

        if request.method == 'POST':
            try:
                # Verificar token
                token = request.form.get('token')
                if not token or session.get('token') != token:
                    print("Error: Token inválido o no coincide")
                    raise ValueError("Token inválido o no coincide")
                session.pop('token', None)

                # Validar campos obligatorios
                required_fields = []

                missing_fields = [field for field in required_fields if field not in request.form or not request.form[field].strip()]
                if missing_fields:
                    print(f"Campos faltantes: {missing_fields}")
                    return redirect(url_for('Altaempleados'))

                # Validar campos numéricos
                try:
                    sueldo_imss = float(request.form['SUELDO_IMSS'].replace('$', '').replace(',', ''))
                    monedero = float(request.form['MONEDERO'].replace('$', '').replace(',', ''))
                    nomina = float(request.form['NOMINA'].replace('$', '').replace(',', ''))
                except ValueError as ve:
                    print("Error en campos numéricos")
                    return redirect(url_for('Altaempleados'))

                # Crear objeto empleado
                empleado = Empleados(
                    id=0,
                    nombre=request.form['NOMBRE'],
                    apellido=request.form['APELLIDO'],
                    id_empresa=request.form['EMPRESA'],
                    puesto=request.form['PUESTO'],
                    tipo_empleado=request.form['TIPO_EMPLEADO'],
                    tipo_nomina=request.form['TIPO_NOMINA'],
                    sueldo_imss=sueldo_imss,
                    monedero=monedero,
                    nomina=nomina,
                    banco=request.form['BANCO'],
                    numero_cuenta=request.form['NUM_CUENTA'],
                    clabe=request.form['CLABE'],
                    alta_empleado=request.form.get('ALTA_EMPLEADO', ''),
                    baja_empleado=None,
                    fecha_registro=request.form.get('FECHA_REGISTRO', ''),
                    is_blocked=0,
                    categoria=request.form['CATEGORIA'],
                    no_imss=request.form['NO_IMSS'],
                    curp=request.form['CURP'],
                    ine=request.form.get('INE', ''),
                    rfc=request.form['RFC'],
                    cedula_profesional=request.form.get('CEDULA_PROFESIONAL', ''),
                    estado_civil=request.form['ESTADO_CIVIL'],
                    fecha_nacimiento=request.form['FECHA_NACIMIENTO'],
                    telefono_contacto=request.form['TELEFONO_CONTACTO'],
                    domicilio=request.form['DOMICILIO'],
                    tope_horas_extra=request.form['TOPE_HORAS_EXTRA'],
                    foto_base64=request.form.get('FOTO_BASE64', ''),
                    tipo_sangre=request.form['TIPO_SANGRE'],
                    lugar_nacimiento=request.form['LUGAR_NACIMIENTO'],
                    sexo=request.form['SEXO'],
                    calle=request.form.get('CALLE', ''),
                    manzana=request.form.get('MANZANA', ''),
                    lote=request.form.get('LOTE', ''),
                    numero_exterior=request.form.get('NUMERO_EXTERIOR', ''),
                    numero_interior=request.form.get('NUMERO_INTERIOR', ''),
                    colonia=request.form['COLONIA'],
                    codigo_postal=request.form['CODIGO_POSTAL'],
                    estado=request.form['ESTADO'],
                    telefono_domicilio=request.form.get('TELEFONO_DOMICILIO', ''),
                    cuenta_correo=request.form.get('CUENTA_CORREO', ''),
                    salario_diario_integrado=request.form.get('SALARIO_DIARIO_INTEGRADO', 0.0),
                    numero_credito_infonavit=request.form.get('NUMERO_CREDITO_INFONAVIT', ''),
                    tipo_descuento_infonavit=request.form.get('TIPO_DESCUENTO_INFONAVIT', ''),
                    factor_infonavit=request.form.get('FACTOR_INFONAVIT', 0.0),
                    fecha_ingreso=request.form.get('FECHA_INGRESO', ''),
                    turno=request.form.get('TURNO', ''),
                    tipo_contrato=request.form.get('TIPO_CONTRATO', ''),
                    contacto_accidente=request.form.get('CONTACTO_ACCIDENTE', ''),
                    alergias=request.form.get('ALERGIAS', ''),
                    enfermedades_controladas=request.form.get('ENFERMEDADES_CONTROLADAS', ''),
                    edificio=request.form.get('EDIFICIO', ''),
                    alcaldia=request.form.get('ALCALDIA', ''),
                    municipio=request.form.get('MUNICIPIO', ''),
                    registro_patronal=request.form.get('REGISTRO_PATRONAL', ''),
                    cuenta=request.form.get('CUENTA', ''),
                    contratable=request.form.get('CONTRATABLE', None),
                    observaciones=request.form.get('OBSERVACIONES', '')
                )

                # Insertar en la base de datos
                ModelEmpleado.alta_empleado(db, empleado)
                return redirect(url_for('empleados'))

            except KeyError as ke:
                print(f"Error de campo faltante: {str(ke)}")
                return redirect(url_for('Altaempleados'))

            except Exception as ex:
                print(f"Error en alta empleado: {str(ex)}")
                return redirect(url_for('Altaempleados'))

        else:
            session['token'] = str(uuid.uuid4())
            return render_template('altaEmpleados.html', registosPatronales=registosPatronales, empresas=empresas, bancos=bancos, puesto=puesto, token=session['token'])

    except Exception as e:
        print(f"Error en la ruta /Altaempleados: {str(e)}")
        return redirect(url_for('empleados')) 
    
    
#BLOQUEAR EMPLEADO
@app.route('/block_empleado/<int:id>', methods=['POST'])
def block_empleado(id):
    if request.method == 'POST':
        try:
       
            ModelEmpleado.change_status(db, id, True)
            return redirect(url_for('empleados'))
        except Exception as e:
            return jsonify({'error': str(e)}), 500

#DESBLOQUEA EMPLEADO 
@app.route('/unblock_empleado/<int:id>', methods=['POST'])
def unblock_empleado(id):

    if request.method == 'POST':
        try:
          
            ModelEmpleado.change_status(db, id, False)
            return redirect(url_for('empleados'))
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    

@app.route('/edit_empleado/<int:id>', methods=['GET', 'POST'])
def edit_empleado(id):
    try:
        # Obtener datos iniciales
        puesto = ModelPuesto.get_all_puestos_no_block(db)
        bancos = ModelBanco.get_all_bancos_no_block(db)
        bancos_data = [{"id": banco.id, "nombre": banco.nombre} for banco in bancos]
        empresas = ModelEmpresas.get_empresas_not_block(db)
        registosPatronales = ModelRegistroPatronal.get_all_registros_patronales(db)
        empleado = ModelEmpleado.get_empleado_by_id(db, id)

        if request.method == 'POST':
            try:
                # Validar campos obligatorios
                required_fields = []  # Define aquí los campos requeridos
                missing_fields = [field for field in required_fields if field not in request.form or not request.form[field].strip()]
                if missing_fields:
                    return redirect(url_for('edit_empleado', id=id))

                # Validar campos numéricos
                try:
                    sueldo_imss = float(request.form['SUELDO_IMSS'].replace('$', '').replace(',', ''))
                    monedero = float(request.form['MONEDERO'].replace('$', '').replace(',', ''))
                    nomina = float(request.form['NOMINA'].replace('$', '').replace(',', ''))
                except ValueError:
                    return redirect(url_for('edit_empleado', id=id))

                # Actualizar objeto empleado
                empleado = Empleados(
                    id=id,
                    nombre=request.form['NOMBRE'],
                    apellido=request.form['APELLIDO'],
                    id_empresa=request.form['EMPRESA'],
                    puesto=request.form['PUESTO'],
                    tipo_empleado=request.form['TIPO_EMPLEADO'],
                    tipo_nomina=request.form['TIPO_NOMINA'],
                    sueldo_imss=sueldo_imss,
                    monedero=monedero,
                    nomina=nomina,
                    banco=request.form['BANCO'],
                    numero_cuenta=request.form['NUM_CUENTA'],
                    clabe=request.form['CLABE'],
                    alta_empleado=request.form.get('ALTA_EMPLEADO', ''),
                    baja_empleado=None,
                    fecha_registro=request.form.get('FECHA_REGISTRO', ''),
                    is_blocked=0,
                    categoria=request.form['CATEGORIA'],
                    no_imss=request.form['NO_IMSS'],
                    curp=request.form['CURP'],
                    ine=request.form.get('INE', ''),
                    rfc=request.form['RFC'],
                    cedula_profesional=request.form.get('CEDULA_PROFESIONAL', ''),
                    estado_civil=request.form['ESTADO_CIVIL'],
                    fecha_nacimiento=request.form['FECHA_NACIMIENTO'],
                    telefono_contacto=request.form['TELEFONO_CONTACTO'],
                    domicilio=request.form['DOMICILIO'],
                    tope_horas_extra=request.form['TOPE_HORAS_EXTRA'],
                    foto_base64=request.form.get('FOTO_BASE64', ''),
                    tipo_sangre=request.form['TIPO_SANGRE'],
                    lugar_nacimiento=request.form['LUGAR_NACIMIENTO'],
                    sexo=request.form['SEXO'],
                    calle=request.form.get('CALLE', ''),
                    manzana=request.form.get('MANZANA', ''),
                    lote=request.form.get('LOTE', ''),
                    numero_exterior=request.form.get('NUMERO_EXTERIOR', ''),
                    numero_interior=request.form.get('NUMERO_INTERIOR', ''),
                    colonia=request.form['COLONIA'],
                    codigo_postal=request.form['CODIGO_POSTAL'],
                    estado=request.form['ESTADO'],
                    telefono_domicilio=request.form.get('TELEFONO_DOMICILIO', ''),
                    cuenta_correo=request.form.get('CUENTA_CORREO', ''),
                    salario_diario_integrado=request.form.get('SALARIO_DIARIO_INTEGRADO', 0.0),
                    numero_credito_infonavit=request.form.get('NUMERO_CREDITO_INFONAVIT', ''),
                    tipo_descuento_infonavit=request.form.get('TIPO_DESCUENTO_INFONAVIT', ''),
                    factor_infonavit=request.form.get('FACTOR_INFONAVIT', 0.0),
                    fecha_ingreso=request.form.get('FECHA_INGRESO', ''),
                    turno=request.form.get('TURNO', ''),
                    tipo_contrato=request.form.get('TIPO_CONTRATO', ''),
                    contacto_accidente=request.form.get('CONTACTO_ACCIDENTE', ''),
                    alergias=request.form.get('ALERGIAS', ''),
                    enfermedades_controladas=request.form.get('ENFERMEDADES_CONTROLADAS', ''),
                    edificio=request.form.get('EDIFICIO', ''),
                    alcaldia=request.form.get('ALCALDIA', ''),
                    municipio=request.form.get('MUNICIPIO', ''),
                    registro_patronal=request.form.get('REGISTRO_PATRONAL', ''),
                    cuenta=request.form.get('CUENTA', ''),
                    contratable=request.form.get('CONTRATABLE', None),
                    observaciones=request.form.get('OBSERVACIONES', '')
                )

                # Actualizar en la base de datos
                ModelEmpleado.update_empleado(db, empleado)

                # Imprimir datos del empleado para depuración
                print(empleado)

                return redirect(url_for('empleados'))

            except Exception as ex:
                return redirect(url_for('edit_empleado', id=id))

        # Renderizar la plantilla con los registros patronales y los campos adicionales
        return render_template(
            'edit_empleado.html',
            empleado=empleado,
            empresas=empresas,
            bancos=bancos_data,
            puesto=puesto,
            registosPatronales=registosPatronales
        )

    except Exception as e:
        print(f"Error en la ruta /edit_empleado: {str(e)}")
        return redirect(url_for('empleados'))

"""
FIN EMPLEADOS
"""

"""
ALTA MONEDERO
"""

@app.route('/Monederos', methods=['GET', 'POST'])
def monederos():
    try:
        if request.method == 'POST':
            # Verifica si el token en la sesión coincide con el del formulario
            token = request.form.get('token')
            if token and session.get('token') == token:
                # Elimina el token de la sesión para evitar reenvíos duplicados
                session.pop('token', None)

                banco = request.form.get('BANCO', '')
                estatus = request.form.get('ESTATUS', '')

                # Filtra monederos según los parámetros enviados en el formulario
                monederos = ModelMonedero.filter_monedero(db, banco, estatus)

                # Imprimir los monederos en la consola
                for monedero in monederos:
                    print(monedero)  # Esto invoca automáticamente el método __str__

                # Obtener todos los bancos para pasarlos al formulario
                bancos = ModelBanco.get_all_bancos_no_block(db)

                # Renderizamos la plantilla pasando la lista de monederos y bancos
                return render_template('monedero.html', monederos=monederos, bancos=bancos, token=session.get('token'))

            else:
                return redirect(url_for('monederos'))
        
        else:
            # Genera un token único y lo almacena en la sesión
            session['token'] = str(uuid.uuid4())

            # Obtener los bancos para pasarlos al formulario
            bancos = ModelBanco.get_all_bancos_no_block(db)

            # Obtener todos los monederos de la base de datos
            monederos = ModelMonedero.get_all_monedero(db)

            # Imprimir los monederos en la consola
            for monedero in monederos:
                print(monedero)  # Esto invoca automáticamente el método __str__

            # Renderizamos la plantilla pasando la lista de monederos
            return render_template('monedero.html', monederos=monederos, bancos=bancos, token=session['token'])

    except Exception as ex:
        print(f"Error en la ruta /Monederos: {str(ex)}")
        raise Exception(f"Error: {str(ex)}")


@app.route('/AltaMonedero', methods=['POST'])
def alta_monedero():
    try:
        # Verifica si el token en la sesión coincide con el del formulario
        token = request.form.get('token')
        if token and session.get('token') == token:
            # Elimina el token de la sesión para evitar reenvíos duplicados
            session.pop('token', None)

            # Obtener los datos del formulario
            banco = request.form.get('BANCO_name')
            id_banco = request.form.get('BANCO')
            numero_tarjeta = request.form.get('NUMERO_TARJETA', '')
            estatus = request.form.get('ESTATUS')

            # Crear el objeto Monedero
            monedero = Monedero(
                banco=banco,
                id_banco=id_banco, 
                numero_tarjeta=numero_tarjeta, 
                estatus=estatus
            )

            # Usar el modelo para guardar el monedero en la base de datos
            ModelMonedero.alta_monedero(db, monedero)

            # Redirigir a la página de Monederos con un mensaje de éxito
            return redirect(url_for('monederos'))
        
        else:
            return redirect(url_for('monederos'))

    except Exception as ex:
        return f"Error al procesar el formulario: {str(ex)}"

@app.route('/filtrarMonederos', methods=['GET', 'POST'])
def filtrar_monedero():
    try:
        # Obtener los parámetros de la solicitud
        banco = request.args.get('banco', None)
        estatus = request.args.get('estatus', None)
        numero_tarjeta = request.args.get('numero_tarjeta', None)

        # Llamar al modelo con los filtros proporcionados
        monederos = ModelMonedero.filter_monedero(
            db,
            banco=banco,
            estatus=estatus,
            numero_tarjeta=numero_tarjeta
        )

        # Renderizar la plantilla con los resultados filtrados
        return render_template(
            'monedero.html',
            monederos=monederos,
            bancos=ModelBanco.get_all_bancos(db)
        )
    except Exception as ex:
        # Manejar errores y devolver un mensaje de error
        return f"Error al filtrar monederos: {str(ex)}"


@app.route('/EditarMonedero', methods=['POST'])
def editar_monedero():
    try:
        print("Datos recibidos:", request.form.to_dict())  # Agrega esto para ver los datos en la consola
        # Validar el token de la sesión
        token = request.form.get('token')
        if not token or session.get('token') != token:
            return "Token inválido o ausente.", 400

        session.pop('token', None)  # Eliminar el token después de validar

        # Validar y obtener los datos
        id_monedero = request.form.get('id_monedero')
        banco = request.form.get('BANCO_name')
        id_banco = request.form.get('BANCO')
        numero_tarjeta = request.form.get('NUMERO_TARJETA')
        estatus = request.form.get('ESTATUS')

        if not all([id_monedero, banco, id_banco, numero_tarjeta, estatus]):
            print("Error: datos faltantes.")
            return "Faltan datos obligatorios.", 400

        # Crear el objeto Monedero y actualizar
        monedero = Monedero(
            id=id_monedero,
            banco=banco,
            id_banco=id_banco,
            numero_tarjeta=numero_tarjeta,
            estatus=estatus
        )
        ModelMonedero.update_monedero(db, monedero)

        return redirect(url_for('monederos'))
    except Exception as ex:
        print("Error al actualizar el monedero:", str(ex))  # Agrega esto para depuración
        return f"No se pudo actualizar el monedero. Verifica los datos e intenta nuevamente.", 500

@app.route('/BuscarMonederos', methods=['GET'])
def buscar_monedero():
    numero_tarjeta = request.args.get('numero_tarjeta', '')
    try:
        # Si el input está vacío, devuelve todos los monederos
        if not numero_tarjeta:
            monederos = ModelMonedero.get_all_monedero(db)
        else:
            # Si hay un número de tarjeta, filtra por ese número
            monederos = ModelMonedero.filter_monedero(db, numero_tarjeta=numero_tarjeta)

        # Construye la respuesta incluyendo si están asignados o no
        return jsonify([{
            'id': monedero.id,
            'banco': monedero.banco,
            'numero_tarjeta': monedero.numero_tarjeta,
            'id_empleado': monedero.id_empleado,  # Indica si está asignado a un empleado
        } for monedero in monederos])
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
@app.route('/BuscarMonederosDisponibles', methods=['GET'])
def buscar_monedero_disponibles():
    numero_tarjeta = request.args.get('numero_tarjeta', '')
    try:
        # Si el input está vacío, devuelve todos los monederos
        if not numero_tarjeta:
            monederos = ModelMonedero.get_monedero_disponibles(db)
        else:
            # Si hay un número de tarjeta, filtra por ese número
            monederos = ModelMonedero.filter_monedero_disponible(db, numero_tarjeta=numero_tarjeta)

        # Construye la respuesta incluyendo si están asignados o no
        return jsonify([{
            'id': monedero.id,
            'banco': monedero.banco,
            'numero_tarjeta': monedero.numero_tarjeta,
            'id_empleado': monedero.id_empleado,  # Indica si está asignado a un empleado
        } for monedero in monederos])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/AsignarMonedero', methods=['POST'])
def asignar_monedero():
    empleado_id = request.form.get('empleado_id')
    monedero_id = request.form.get('monedero_id')

    # Validar datos del formulario
    if not empleado_id or not monedero_id:
        app.logger.error("Faltan datos en el formulario: empleado_id=%s, monedero_id=%s", empleado_id, monedero_id)
        flash('Faltan datos necesarios para asignar el monedero.', 'error')
        return redirect(url_for('empleados'))

    try:
        # Verificar que el monedero exista
        monedero = ModelMonedero.get_monedero_by_id(db, monedero_id)
        if not monedero:
            app.logger.error("Monedero no encontrado: monedero_id=%s", monedero_id)
            flash('Monedero no encontrado.', 'error')
            return redirect(url_for('empleados'))

        # Verificar si el monedero ya está asignado
        if monedero.id_empleado:
            app.logger.warning("Monedero ya asignado: monedero_id=%s, id_empleado=%s", monedero_id, monedero.id_empleado)
            flash(f'El monedero ya está asignado al empleado con ID {monedero.id_empleado}.', 'error')
            return redirect(url_for('empleados'))

        # Actualizar el monedero para asociarlo con el empleado
        ModelMonedero.update_id_empleado(db, monedero_id, empleado_id)
        app.logger.info("Monedero asignado correctamente: monedero_id=%s, empleado_id=%s", monedero_id, empleado_id)

        # Actualizar el campo id_monedero en la tabla EMPLEADOS
        ModelEmpleado.update_id_monedero(db, empleado_id, monedero_id)
        app.logger.info("Empleado actualizado con id_monedero: empleado_id=%s, monedero_id=%s", empleado_id, monedero_id)

        flash('Monedero asignado correctamente y datos actualizados.', 'success')

    except Exception as e:
        app.logger.error("Error al asignar monedero: %s", str(e))
        flash(f'Error al asignar el monedero: {str(e)}', 'error')

    return redirect(url_for('empleados'))

@app.route('/DesasignarMonedero', methods=['POST'])
def desasignar_monedero():
    empleado_id = request.form.get('empleado_id')
    monedero_id = request.form.get('monedero_id')

    # Validar los datos del formulario
    if not empleado_id or not monedero_id:
        flash('Faltan datos necesarios para desasignar el monedero.', 'error')
        return redirect(url_for('empleados'))

    try:
        # Actualizar el monedero para desasociarlo del empleado
        ModelMonedero.update_id_empleado(db, monedero_id, None)
        app.logger.info("Monedero desasociado correctamente: monedero_id=%s", monedero_id)

        # Actualizar el empleado para eliminar la referencia al monedero
        ModelEmpleado.update_id_monedero(db, empleado_id, None)
        app.logger.info("Empleado actualizado: empleado_id=%s, id_monedero=NULL", empleado_id)

        flash('Monedero desasignado correctamente.', 'success')
    except Exception as e:
        app.logger.error("Error al desasignar monedero: %s", str(e))
        flash(f'Error al desasignar el monedero: {str(e)}', 'error')

    return redirect(url_for('empleados'))



@app.route('/cambiarEstatus', methods=['POST'])
def cambiar_estatus():
    try:
        # Obtener datos del formulario
        id_monedero = request.form.get('id_monedero')
        nuevo_estatus = request.form.get('estatus')

        if not id_monedero or nuevo_estatus is None:
            return "Faltan datos requeridos", 400

        # Cambiar el estatus del monedero usando el modelo
        ModelMonedero.change_status(db, id=id_monedero, estatus=nuevo_estatus)

        # Redirigir a la página de Monederos después de actualizar
        return redirect(url_for('monederos'))

    except Exception as ex:
        return f"Error al cambiar el estatus del monedero: {str(ex)}"

@app.route('/getMonedero/<int:id>', methods=['GET'])
def get_monedero_by_id(id):
    try:
        monedero = ModelMonedero.get_monedero_by_id(db, id)
        if not monedero:
            return "Monedero no encontrado", 404

        # Respuesta con los datos del monedero
        return jsonify({
            "id": monedero.id,
            "banco": monedero.banco,
            "numero_tarjeta": monedero.numero_tarjeta,  # Verifica este valor
            "id_banco": monedero.id_banco,
            "estatus": monedero.estatus,  # Verifica que el estatus sea '0' o '1'
            "id_empleado": monedero.id_empleado
        })
    except Exception as ex:
        return f"Error al obtener el monedero: {str(ex)}", 500

"""
fIN MONEDERO
"""



"""
INICIA PUESTOS
"""
#PUESTO
@app.route('/Puestos', methods=['GET','POST'])
def puestos():
    try:
        
        categorias = ModelCategoria.get_all_categorias(db)
        
        if request.method == 'POST':
            # Verifica si el token en la sesión coincide con el del formulario
            token = request.form.get('token')
            if token and session.get('token') == token:
                # Elimina el token de la sesión para evitar reenvíos duplicados
                session.pop('token', None)
                
                puesto = request.form['PUESTO']
                estatus = request.form['ESTATUS']
                puestos = ModelPuesto.filter_puesto(db,puesto,estatus)
                
                return render_template('puestos.html', categorias=categorias,puesto=puestos)
                
            else:
                return redirect(url_for('puestos'))
        
        else:
            # Genera un token único y lo almacena en la sesión
            session['token'] = str(uuid.uuid4())
            puestos = ModelPuesto.get_all_puestos(db)
            
            return render_template('puestos.html', categorias=categorias,puesto=puestos,token=session['token'])
    except Exception as e:
        raise e
    

#ALTA PUESTO
@app.route('/Altapuesto', methods=['POST'])
def altapuesto():

    if request.method == 'POST':
           
            puesto = Puesto(
                id=0,
                puesto=request.form['PUESTO'],
                sueldo_base= convertir_a_float(request.form['SUELDO_BASE']),
                sueldo_tarjeta=convertir_a_float(request.form['SUELDO_TARJETA']),
                horas_extras=convertir_a_float(request.form['HORA_EXTRA']),
                categoria=request.form['CATEGORIA']
            )

            ModelPuesto.alta_puesto(db,puesto)
            
            return redirect(url_for('puestos'))

# EDITAR PUESTO
@app.route('/edit_puesto/<int:id>', methods=['POST'])
def edit_puesto(id):
    if request.method == 'POST':
        puesto = Puesto(id=id,
            puesto=request.form['PUESTO_EDIT'],
            sueldo_base=convertir_a_float(request.form['SUELDO_BASE_EDIT']),
            sueldo_tarjeta=convertir_a_float(request.form['SUELDO_TARJETA_EDIT']),
            horas_extras=convertir_a_float(request.form['HORA_EXTRA_EDIT']),
            categoria=request.form['CATEGORIA_EDIT']
        )
        
        ModelPuesto.update_puesto(db,puesto)
        return redirect(url_for('puestos'))

    

# BLOQUEAR PUESTOS
@app.route('/block_puesto/<int:id>', methods=['POST'])
def block_puesto(id):
    if request.method == 'POST':
        try:
       
            ModelPuesto.change_status(db, id, True)
            return redirect(url_for('puestos'))
        except Exception as e:
            return jsonify({'error': str(e)}), 500

#DESBLOQUEA PUESTOS
@app.route('/unblock_puesto/<int:id>', methods=['POST'])
def unblock_puesto(id):
    if request.method == 'POST':
        try:
          
            ModelPuesto.change_status(db, id, False)
            return redirect(url_for('puestos'))
        except Exception as e:
            return jsonify({'error': str(e)}), 500                
       


"""
FIN PUESTOS
"""


"""
    INICIA PROVEEDORES
"""

#ALTA PROVEEDORES
@app.route('/Altaproveedores', methods=['GET','POST'])
def altaproveedores():
    if request.method == 'POST':
        
        # Verifica si el token en la sesión coincide con el del formulario
        token = request.form.get('token')
        if token and session.get('token') == token:
            # Elimina el token de la sesión para evitar reenvíos duplicados
            session.pop('token', None)
            
        
            new_proveedor = Proveedores(
                id=request.form['ID_PROVEEDOR'], 
                id_empresa = current_user.id_empresa,
                razon_social=request.form['RAZON_SOCIAL'],
                regimen_fiscal_id=request.form['REGIMEN_FISCAL'],
                tipo_id=request.form['TIPO'],
                rfc=request.form['RFC'],
                pais=request.form['PAIS'],
                estado=request.form['ESTADO'],
                cp=request.form['CP'],
                municipio=request.form['MUNICIPIO'],
                colonia=request.form['COLONIA'],
                calle=request.form['CALLE'],
                numero_exterior=request.form['NUMERO_EXTERIOR'],
                numero_interior=request.form['NUMERO_INTERIOR'],
                usuario=current_user.id
            )

            id_proveedor = ModelProveedores.crear_proveedor(db,new_proveedor)

            especialidades =  request.form.getlist('ID_ESPECIALIDADES[]')
            for e in range(len(especialidades)):
                nueva_especialida = EspecialidadesProveedor(
                    id=0,
                    id_especialidad=especialidades[e],
                    id_proveedor=id_proveedor,
                )

                ModelEspecialidadesProveedor.crear_especialidad_proveedor(db,nueva_especialida)

            nombres_contactos = request.form.getlist('NOMBRE_CONTACTO[]')
            apellidos_coctactos = request.form.getlist('APELLIDO_CONTACTO[]')
            telefonos_contacto = request.form.getlist('TELEFONO_CONTACTO[]')
            correos_contacto = request.form.getlist('CORREO_CONTACTO[]')
            puestos_contacto = request.form.getlist('PUESTO_CONTACTO[]')

            
            for i in range(len(nombres_contactos)):
                nuevo_contacto = ContactosProveedores(
                    id = 0,
                    id_proveedor= id_proveedor,
                    nombre= nombres_contactos[i],
                    apellido= apellidos_coctactos[i],
                    puesto=puestos_contacto[i],
                    telefono=telefonos_contacto[i],
                    correo=correos_contacto[i]
                )

                ModelContactosProveedor.new_contacto_proveedor(db,nuevo_contacto)
            

            bancos = request.form.getlist('BANCO[]')
            numeros_cuentas = request.form.getlist('NO_CUENTA[]')
            clabes = request.form.getlist('CLABE[]')

            for b in range(len(bancos)):
                nueva_cuenta = CuentasProveedor(
                    id=0,
                    id_proveedor=id_proveedor,
                    id_banco=bancos[b],
                    numero_cuenta=numeros_cuentas[b],
                    clabe=clabes[b]
                )

                ModelCuentasProveedor.new_cuenta_proveedor(db,nueva_cuenta)

            return redirect(url_for('proveedores'))
        
        else:
            return redirect(url_for('proveedores'))
            
        
    else:   
        # Genera un token único y lo almacena en la sesión
        session['token'] = str(uuid.uuid4())
        ultimo_proveedor = ModelProveedores.get_proveedor_last_by_id(db)

        if ultimo_proveedor is None:
                ultimo_proveedor = 1
                
        else:
            ultimo_proveedor =  ultimo_proveedor + 1

        return render_template('altaProveedores.html',siguiente=ultimo_proveedor,token=session['token'])

#VALIDAR RFC DE PROVEEDORES
@app.route('/api/validar_rfc_proveedores/', methods=['GET'])
def validar_rfc_proveedores():
    
    try:
        rfc = request.args.get('rfc', '')

        proveedores = ModelProveedores.get_all_proveedores_not_blocked(db)

        existe_rfc = False

        for p in proveedores:
            if p.rfc == rfc:
                existe_rfc = True
              
            
        return jsonify(existe_rfc), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500

#PROVEEDORES
@app.route('/Proveedores', methods=['GET','POST'])
def proveedores ():
    if request.method == 'POST':
        proveedor  = request.form.get('PROVEEDOR', '')
        regimen_fiscal = request.form.get('REGIMEN_FISCAL', '')
        rfc = request.form.get('RFC', '')
        estatus = request.form.get('ESTATUS', '')

        proveedores = ModelProveedores.filter_proveedores(db, proveedor, regimen_fiscal, rfc,estatus)
        
        return render_template('proveedores.html', proveedores=proveedores)

    
    else:
        # Genera un token único y lo almacena en la sesión
        session['token'] = str(uuid.uuid4())
        proveedores = ModelProveedores.get_all_proveedores(db)
        return render_template('proveedores.html', proveedores=proveedores,token=session['token'])

#EDITAR PROVEEDOR
@app.route('/edit_proveedor/<int:id>', methods=['GET', 'POST'])
def edit_proveedor(id):
    if request.method == 'POST':
        
        # Verifica si el token en la sesión coincide con el del formulario
        token = request.form.get('token')
        if token and session.get('token') == token:
            # Elimina el token de la sesión para evitar reenvíos duplicados
            session.pop('token', None)
            
            new_proveedor = Proveedores(
                id=id, 
                id_empresa = current_user.id_empresa,
                razon_social=request.form['RAZON_SOCIAL'],
                regimen_fiscal_id=request.form['REGIMEN_FISCAL'],
                tipo_id=request.form['TIPO'],
                rfc=request.form['RFC'],
                pais=request.form['PAIS'],
                estado=request.form['ESTADO'],
                cp=request.form['CP'],
                municipio=request.form['MUNICIPIO'],
                colonia=request.form['COLONIA'],
                calle=request.form['CALLE'],
                numero_exterior=request.form['NO_EXTERIOR'],
                numero_interior=request.form['NO_INTERIOR']
            )

            ModelProveedores.update_proveedor(db,new_proveedor)

            #ELIMINAR ESPECIALIDADES ASIGNADAS AL PROVEEDOR
            ModelEspecialidadesProveedor.delete_especialidades(db,id)

            especialidades =  request.form.getlist('ID_ESPECIALIDADES[]')
            for e in range(len(especialidades)):
                nueva_especialida = EspecialidadesProveedor(
                    id=0,
                    id_especialidad=especialidades[e],
                    id_proveedor=new_proveedor.id,
                )

                ModelEspecialidadesProveedor.crear_especialidad_proveedor(db,nueva_especialida)

            nombres_contactos = request.form.getlist('NOMBRE_CONTACTO[]')
            apellidos_coctactos = request.form.getlist('APELLIDO_CONTACTO[]')
            telefonos_contacto = request.form.getlist('TELEFONO_CONTACTO[]')
            correos_contacto = request.form.getlist('CORREO_CONTACTO[]')
            puestos_contacto = request.form.getlist('PUESTO_CONTACTO[]')

            # ELIMINAR CONTACTOS PROVEEDOR
            ModelContactosProveedor.delete_contactos_proveedor(db,id)

            for i in range(len(nombres_contactos)):
                nuevo_contacto = ContactosProveedores(
                    id = 0,
                    id_proveedor= new_proveedor.id,
                    nombre= nombres_contactos[i],
                    apellido= apellidos_coctactos[i],
                    puesto=puestos_contacto[i],
                    telefono=telefonos_contacto[i],
                    correo=correos_contacto[i]
                )

                ModelContactosProveedor.new_contacto_proveedor(db,nuevo_contacto)
            
            # ELIMINAR CUENTAS PROVEEDOR
            ModelCuentasProveedor.delete_cuentas_proveedor(db,id)

            bancos = request.form.getlist('BANCO[]')
            numeros_cuentas = request.form.getlist('NO_CUENTA[]')
            clabes = request.form.getlist('CLABE[]')

            for b in range(len(bancos)):
                nueva_cuenta = CuentasProveedor(
                    id=0,
                    id_proveedor=new_proveedor.id,
                    id_banco=bancos[b],
                    numero_cuenta=numeros_cuentas[b],
                    clabe=clabes[b]
                )

                ModelCuentasProveedor.new_cuenta_proveedor(db,nueva_cuenta)

            return redirect(url_for('proveedores'))
        
        else:
            return redirect(url_for('proveedores'))
            
            
    else:
        proveedor = ModelProveedores.get_proveedor_by_id(db,id)
        especialidades = ModelEspecialidadesProveedor.get_especialidades_by_proveedor(db,id)
        contactos = ModelContactosProveedor.get_contactos_by_proveedor(db,id)
        cuentas = ModelCuentasProveedor.get_cuentas_by_proeveedor(db,id)
        bancos = ModelBanco.get_all_bancos(db)
        return render_template('edit_proveedor.html', proveedor=proveedor,especialidades=especialidades,contactos=contactos,cuentas=cuentas,bancos=bancos,token=session['token'])


# BLOQUEAR PROVEEDORES
@app.route('/block_proveedores/<int:id>', methods=['POST'])
def block_proveedores(id):
    if request.method == 'POST':
        try:
       
            ModelProveedores.change_status(db, id, True)
            return redirect(url_for('proveedores'))
        except Exception as e:
            return jsonify({'error': str(e)}), 500

#DESBLOQUEA PROVEEDORES
@app.route('/unblock_proveedores/<int:id>', methods=['POST'])
def unblock_proveedores(id):
    if request.method == 'POST':
        try:
            ModelProveedores.change_status(db, id, False)
            return redirect(url_for('proveedores'))
        except Exception as e:
            return jsonify({'error': str(e)}), 500                
   
"""
    FIN PROVEEDORES
"""


# ESPECIALIDADES
@app.route('/Especialidades',methods=['GET','POST'])
def especialidades():
    if request.method == 'POST':
        new_especialidad = Especialidades(
            id=0,
            nombre=request.form['ESPECIALIDAD']
        )
        
        ModelEspecialidades.crearEspecialidad(db,new_especialidad)
        return redirect(url_for('especialidades'))
    
    especialidades = ModelEspecialidades.get_all_especialidades(db)
    return render_template('especialidades.html',especialidades=especialidades)

## FILTRO ESPECIALIDADES   
@app.route('/fitro_especialidad', methods=['POST'])
@login_required
def fitro_especialidad():
    especialidad = request.form.get('especialidad', '')
    estatus = request.form.get('estatus', '')

    try:
        especialidad = str(especialidad)
        especialidades = ModelEspecialidades.filter_especialidad(db,especialidad, estatus)
       
        if especialidades:
            return render_template('especialidades.html', especialidades=especialidades)
        else:
            return redirect(url_for('especialidades'))
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500



# EDITAR ESPECIALIDADES
@app.route('/edit_especialidad/<int:id>', methods=['POST'])
def editar_especialidad(id):
    if request.method == 'POST':
       
        especialidad = Especialidades(
            id=id,
            nombre=request.form['ESPECIALIDAD']
        )
    
    ModelEspecialidades.update_especialidad(db,especialidad)
    return redirect(url_for('especialidades'))

# BLOQUEAR ESPECIALIDADES
@app.route('/block_especialidades/<int:id>', methods=['POST'])
def block_especialidades(id):
    if request.method == 'POST':
        try:
       
            ModelEspecialidades.change_status(db, id, True)
            return redirect(url_for('especialidades'))
        except Exception as e:
            return jsonify({'error': str(e)}), 500

#DESBLOQUEA ESPECIALIDADES
@app.route('/unblock_especialidades/<int:id>', methods=['POST'])
def unblock_especialidades(id):
    if request.method == 'POST':
        try:
          
            ModelEspecialidades.change_status(db, id, False)
            return redirect(url_for('especialidades'))
        except Exception as e:
            return jsonify({'error': str(e)}), 500                
            

# API LLENAR PORYECTO
@app.route('/api/buscar_proyecto/', methods=['GET'])
def buscar_proyecto():
    try:
        query = request.args.get('query', '')
        proyectos = ModelProyectoObra.get_proyectos_not_block(db)
        
        proyectos_filtradas = []

        for p in proyectos:
            if query.lower() in p.nombre_proyecto.lower():
                # Definir las fechas como cadenas para evitar el error de `datetime.date`
                fecha_inicio = datetime.strptime(str(p.fecha_inicio), '%Y-%m-%d')
                fecha_fin = datetime.strptime(str(p.fecha_fin), '%Y-%m-%d')
                
                # Calcular la diferencia entre las dos fechas
                diferencia = fecha_fin - fecha_inicio
                semanas = round(diferencia.days / 7)
                
                direccion = f"{p.calle} {p.numero_exterior}"
                if p.numero_interior:
                    direccion += f" Int. {p.numero_interior}"
                direccion += f", {p.colonia}, {p.municipio}, {p.estado}, {p.pais}"
                
                director = ModelEmpleado.obtener_director_por_id(db,p.director_proyecto)
                
                proyectos_filtradas.append({
                    'id': p.id,
                    'id_cliente': p.id_cliente,
                    'nombre_proyecto': p.nombre_proyecto,
                    'fecha_inicio': p.fecha_inicio,
                    'fecha_fin': p.fecha_fin,
                    'semanas': semanas,
                    'director_proyecto': director.nombre + ' ' + director.apellido,
                    'id_director': director.id,
                    'direccion': direccion
                })
        
        if proyectos_filtradas:
            return jsonify(proyectos_filtradas), 200
        else:
            return jsonify([]), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500

    
@app.route('/api/get_project_details/<int:id>', methods=['GET'])
def get_project_details(id):
    try:
        proyecto = ModelProyectoObra.get_proyecto_by_id(db, id)
        if proyecto:
            return jsonify({
                'id': proyecto.id,
                'nombre_proyecto': proyecto.nombre_proyecto,
                'lider_proyecto': proyecto.lider_proyecto,
                'director_proyecto': proyecto.director_proyecto
            }), 200
        else:
            return jsonify({'error': 'Proyecto no encontrado'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# API LLENAR FAMILIA
@app.route('/api/llenar_familia/', methods=['GET'])
def get_all_familia():
    try:
        query = request.args.get('query', '')

        # Obtener todas las familias desde el modelo
        familias = ModelFamilias.get_all_familias(db)

        # Filtrar las familias basadas en la query
        familias_filtradas = [
            {'id': e.id, 'familia': e.familia}
            for e in familias
            if query.lower() in e.familia.lower()
        ]

        # Devolver el resultado filtrado como JSON
        return jsonify(familias_filtradas), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
@app.route('/api/filtrar_familia/', methods=['GET'])
def filtrar_familia():
    try:
        # Obtener parámetros de la solicitud
        familia = request.args.get('familia', '').strip()
        estado = request.args.get('estado', '').strip().lower()

        # Validar estado (opcional)
        if estado not in ('', 'activo', 'bloqueado'):
            return jsonify({'error': 'Estado no válido, debe ser "activo" o "bloqueado"'}), 400

        # Filtrar familias usando el método del modelo
        familias = ModelFamilias.filter_familia(db, familia=familia, estado=estado)

        # Construir la respuesta como JSON
        familias_data = [
            {
                'id': f.id,
                'familia': f.familia,
                'is_blocked': f.is_blocked
            }
            for f in familias
        ]

        return jsonify(familias_data), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
# API OBTENER CATEGORIAS
@app.route('/api/obtener_categorias/', methods=['GET'])
def get_all_categorias():
    try:
        query = request.args.get('query', '')

        # Obtener todas las familias desde el modelo
        categorias = ModelCategoria.get_all_categorias(db)

        # Filtrar las familias basadas en la query
        categorias_filtradas = [
            {'id': c.idCategoria, 
            'categoria': c.categoria}
            for c in categorias
            if query.lower() in c.categoria.lower()
        ]

        # Devolver el resultado filtrado como JSON
        return jsonify(categorias_filtradas), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# API CONTRATISTA
@app.route('/api/llenar_contratistas/', methods=['GET'])
def get_all_contratistas():
    try:
        id_especialidad = request.args.get('id_especialidad', None)
        query = request.args.get('query', '')

        if not id_especialidad:
            return jsonify({'error': 'El parámetro id_especialidad es requerido'}), 400

        id_especialidad = int(id_especialidad)  # Asegúrate de que es un entero
      

        contratistas = Modelcontratistas.get_all_contratistas(db, id_especialidad)


        contratistas_filtradas = []
        for e in contratistas:
            if query.lower() in e.razon_social.lower():
                contratistas_filtradas.append({ 
                    'id_especialidad': e.id_especialidad,
                    'id_proveedor': e.id_proveedor,
                    'nombre': e.nombre,
                    'razon_social': e.razon_social
                })
                 
        if contratistas_filtradas:
            return jsonify(contratistas_filtradas), 200
        else:
            return jsonify([]), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# API LENAR PUESTOS
@app.route('/api/llenar_puestos/', methods=['GET'])
def get_puestos_by_category():
    try:
        id_categoria = request.args.get('id_categoria', None)
  

        if not id_categoria:
            return jsonify({'error': 'El parámetro id_especialidad es requerido'}), 400

        id_categoria = int(id_categoria)  # Asegúrate de que es un entero
      

        puestos = ModelPuesto.get_puestos_by_category(db, id_categoria)

        listaPuestosFiltrados = []
        for p in puestos:
            listaPuestosFiltrados.append({ 
                    'id': p.id,
                    'puesto': p.puesto,
                    'sueldo_base': p.sueldo_base,
                    'sueldo_tarjeta': p.sueldo_tarjeta,
                    'horas_extras' : p.horas_extras
                })
                 
        if listaPuestosFiltrados:
            return jsonify(listaPuestosFiltrados), 200
        else:
            return jsonify([]), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# API OPNTER PUESTO POR ID
@app.route('/api/obtener_puesto/', methods=['GET'])
def get_puesto_by_id():
    try:
        # Obtener el parámetro id_puesto de la URL
        id_puesto = request.args.get('id_puesto', None)

        if not id_puesto:
            return jsonify({'error': 'El parámetro id_puesto es requerido'}), 400

        # Convertir a entero (manejar errores si no es válido)
        try:
            id_puesto = int(id_puesto)
        except ValueError:
            return jsonify({'error': 'El parámetro id_puesto debe ser un número entero'}), 400

        # Llamar al modelo para obtener el puesto
        puestos = ModelPuesto.get_puestos_by_id(db, id_puesto)

        if not puestos:  # Verificar si no se encontraron resultados
            return jsonify({'error': f'No se encontró un puesto con id {id_puesto}'}), 404

        # Construir el diccionario de respuesta
        puesto = {
            'id': puestos.id,
            'puesto': puestos.puesto,
            'sueldo_base': puestos.sueldo_base,
            'sueldo_tarjeta': puestos.sueldo_tarjeta,
            'horas_extras': puestos.horas_extras
        }

        # Respuesta exitosa
        return jsonify(puesto), 200

    except Exception as e:
        # Manejo de excepciones generales
        return jsonify({'error': f'Error interno del servidor: {str(e)}'}), 500

# API CLIENTE
@app.route('/api/buscar_cliente/', methods=['GET'])
def buscar_cliente():
    try:
        query = request.args.get('query', '')

        clientes = Modelclientes.get_clientes_not_block(db)

        clientes_filtradas = []

        for e in clientes:
            if query.lower() in e.razon_social.lower():
                clientes_filtradas.append({
                    'id': e.id,
                    'id_empresa': e.id_empresa,
                    'razon_social': e.razon_social,
                    'rfc': e.rfc,
                    'cp': e.cp,
                    'estado': e.estado,
                    'ciudad': e.ciudad,
                    'municipio': e.municipio,
                    'pais': e.pais,
                    'calle': e.calle,
                    'no_exterior': e.no_exterior,
                    'regimen_fiscal_id': e.regimen_fiscal_id,
                    'uso_cfdi_id': e.uso_cfdi_id,
                    'condicion_pago': e.condicion_pago,
                    'forma_pago_id': e.forma_pago_id,
                    'fecha_registro': e.fecha_registro,
                    'usuario_id': 0,
                    'is_blocked': e.is_blocked,
                })

        if clientes_filtradas:
            return jsonify(clientes_filtradas), 200
        else:
            return jsonify([]), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# API CLIENTE
@app.route('/api/buscar_empresa/', methods=['GET'])
def buscar_empresa():
    try:
        query = request.args.get('query', '')

        empresas = ModelEmpresas.get_all_empresas(db)

        empresas_filtradas = []

        for e in empresas:
            if query.lower() in e.razon_social.lower():
                empresas_filtradas.append({
                    'id': e.id,
                    'razon_social': e.razon_social,

                })

        if empresas_filtradas:
            return jsonify(empresas_filtradas), 200
        else:
            return jsonify([]), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500



# API CLIENTE
@app.route('/api/obtener_Cliente/', methods=['GET'])
def obtener_clientes():
    try:
        query = request.args.get('query', '')

        clientes = Modelclientes.get_all_clientes(db)

        clientes_filtradas = []

        for e in clientes:
            empresa = ModelEmpresas.get_empresa_by_id(db,e.id_empresa)
            if query.lower() in e.razon_social.lower():
                clientes_filtradas.append({
                    'id': e.id,
                    'id_empresa': e.id_empresa,
                    'nombre_empresa': empresa.razon_social,
                    'razon_social': e.razon_social,
                    'rfc': e.rfc,
                    'cp': e.cp,
                    'estado': e.estado,
                    'ciudad': e.ciudad,
                    'municipio': e.municipio,
                    'pais': e.pais,
                    'calle': e.calle,
                    'no_exterior': e.no_exterior,
                    'regimen_fiscal_id': e.regimen_fiscal_id,
                    'uso_cfdi_id': e.uso_cfdi_id,
                    'condicion_pago': e.condicion_pago,
                    'forma_pago_id': e.forma_pago_id,
                    'fecha_registro': e.fecha_registro,
                    'usuario_id': 0,
                    'is_blocked': e.is_blocked,
                })

        if clientes_filtradas:
            return jsonify(clientes_filtradas), 200
        else:
            return jsonify([]), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


    


@app.route('/api/obtener_especialidades/', methods=['GET'])
def obtener_especialidades():
    try:
        query = request.args.get('query', '')

        especialidades = ModelEspecialidades.get_all_especialidades_not_block(db)
        especialidades_filtradas = []

        for e in especialidades:
            if query.lower() in e.nombre.lower():
                especialidades_filtradas.append({
                    'id': e.id,
                    'nombre': e.nombre
                })

        if especialidades_filtradas:
            return jsonify(especialidades_filtradas), 200
        else:
            return jsonify([]), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/obtener_puestos/', methods=['GET'])
def obtener_puestos():
    try:
        query = request.args.get('query', '')

        puestos =ModelPuesto.get_all_puestos_no_block(db)
        lista_puestos_filtrados = []

        for p in puestos:
            if query.lower() in e.puesto.lower():
                lista_puestos_filtrados.append({
                    'id':p.id,
                    'puesto':p.puesto,
                    'sueldo_base':p.sueldo_base,
                    'sueldo_tarjeta': p.sueldo_tarjeta,
                    'horas_extra':p.horas_extra,
                    'categoria': p.categoria
                })

        if lista_puestos_filtrados:
            return jsonify(lista_puestos_filtrados), 200
        else:
            return jsonify([]), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/obtener_materiales/', methods=['GET'])
def obtener_materiales():
    try:
        query = request.args.get('query', '')

        especialidades =ModelMateriales.get_all_materiales(db)
        especialidades_filtradas = []

        for e in especialidades:
            if query.lower() in e.descripcion.lower():
                especialidades_filtradas.append({
                    'id': e.id,
                    'nombre': e.descripcion
                })

        if especialidades_filtradas:
            return jsonify(especialidades_filtradas), 200
        else:
            return jsonify([]), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/filtrar_materiales/', methods=['GET'])
def filtrar_materiales():
    try:
        # Obtener parámetros de la solicitud
        descripcion = request.args.get('descripcion', '').strip()
        estado = request.args.get('estado', '').strip().lower()
        familia = request.args.get('familia', '').strip()
        unidad_medida = request.args.get('unidad_medida', '').strip()

        # Validar estado (opcional)
        if estado not in ('', 'activo', 'bloqueado', ''):
            return jsonify({'error': 'Estado no válido, debe ser "activo" o "bloqueado"'}), 400

        # Llamar al método del modelo para filtrar materiales
        materiales = ModelMateriales.filter_material_2(
            db,
            descripcion=descripcion,
            estado=estado,
            familia=familia,
            unidad_medida=unidad_medida
        )

        # Construir la respuesta como JSON
        materiales_data = [
            {
                'id': m.id,
                'descripcion': m.descripcion,
                'familia': m.clave_id,  # Ajustado al campo ID_FAMILIA
                'unidad_medida': m.unidad_medida,
                'is_blocked': m.is_blocked
            }
            for m in materiales
        ]

        return jsonify(materiales_data), 200

    except Exception as e:
        # Registrar el error para depuración
        print(f"Error al filtrar materiales: {str(e)}")
        return jsonify({'error': 'Error interno del servidor al filtrar materiales.'}), 500

@app.route('/api/editar_material/<int:id>', methods=['POST'])
def editar_material(id):
    try:
        # Obtener los datos enviados en la solicitud
        data = request.get_json()  # Asegúrate de que el cliente envíe JSON correctamente
        if not data:
            raise KeyError("Datos vacíos o no enviados correctamente.")

        # Validar campos requeridos
        required_fields = ['id_familia', 'material', 'unidad_medida']
        for field in required_fields:
            if field not in data:
                raise KeyError(f"Falta el campo requerido: {field}")

        # Construir el objeto Materiales
        material_obj = Materiales(
            id=id,
            clave_id=data['id_familia'],  # Ahora asignamos correctamente el id_familia
            descripcion=data['material'],
            unidad_medida=data['unidad_medida'],
            fecha_registro=data.get('fecha_registro', None),  # Si no se envía, será None
            usuario=data.get('usuario_id', None),  # Si no se envía, será None
            is_blocked=data.get('is_blocked', 0)  # Default a 0 si no se envía
        )

        # Llamar al método del modelo para actualizar
        ModelMateriales.update_material_2(db, material_obj)

        return jsonify({'message': 'Material actualizado correctamente.'}), 200

    except KeyError as e:
        # Si falta algún campo clave
        print(f"Error de KeyError: {e}")
        return jsonify({'error': f'Falta un campo clave: {str(e)}'}), 400

    except Exception as e:
        # Si ocurre cualquier otro error
        print(f"Error al editar material: {e}")
        return jsonify({'error': f'Error interno del servidor: {str(e)}'}), 500


#CATEGORIAS
@app.route('/Categorias', methods=['GET', 'POST'])
def categorias():
    if request.method == 'POST':
        categoria = Categoria(idCategoria=0,
                                  categoria=request.form['CATEGORIA'])
        
        ModelCategoria.alta_categoria(db,categoria)
        return redirect(url_for('categorias'))
        
   
    categorias = ModelCategoria.get_all_categorias(db)
  
    return render_template('categorias.html', categorias = categorias)

#BLOCK CATEGORIA
@app.route('/block_categoria/<int:id>', methods=['POST'])
def block_categoria(id):
    if request.method == 'POST':
        try:
            ModelCategoria.change_status(db, id, True)
            return redirect(url_for('categorias'))
        except Exception as e:

            return jsonify({'error': str(e)}), 500

#UNBLOCK CATEGORIA
@app.route('/unblock_categoria/<int:id>', methods=['POST'])
def unblock_categoria(id):
    if request.method == 'POST':
        try:
        
            ModelCategoria.change_status(db, id, False)
            return redirect(url_for('categorias'))
        except Exception as e:
            return jsonify({'error': str(e)}), 500
           

# EDIT CATEGORIA
@app.route('/edit_categoria/<int:id>', methods=['GET', 'POST'])
def edit_categoria(id):
    if request.method == 'POST':
        categoria = Categoria(
            idCategoria=id,
            categoria=request.form['EDIT_CATEGORIA']
        )
        
        ModelCategoria.update_categoria(db, categoria)
        return redirect(url_for('categorias'))
    
    categoria = ModelCategoria.get_categoria_by_id(db, id)
    
    return render_template('edit_categoria.html')

#BANCOS
@app.route('/Bancos', methods=['GET', 'POST'])
def bancos():
    if request.method == 'POST':
        new_banco = Banco(id=0,
                      nombre = request.form['BANCO'],
                      usuario="",
                      )
        
        ModelBanco.newBanco(db,new_banco)
        return redirect(url_for('bancos'))
        
   
    bancos = ModelBanco.get_all_bancos(db)
    return render_template('Bancos.html', bancos = bancos)

# EDITAR BANCOS
@app.route('/edit_banco/<int:id>', methods=['POST'])
def edit_banco(id):
    if request.method == 'POST':
       
        banco = Banco(
            id=id,
            nombre=request.form['BANCO'],
             usuario="",
        )
    
    ModelBanco.update_banco(db,banco)
    return redirect(url_for('bancos'))

# BLOQUEAR BANCO
@app.route('/block_banco/<int:id>', methods=['POST'])
def block_banco(id):
    if request.method == 'POST':
        try:
       
            ModelBanco.change_status(db, id, True)
            return redirect(url_for('bancos'))
        except Exception as e:
            return jsonify({'error': str(e)}), 500

#DESBLOQUEA BANCO
@app.route('/unblock_banco/<int:id>', methods=['POST'])
def unblock_banco(id):
    if request.method == 'POST':
        try:
          
            ModelBanco.change_status(db, id, False)
            return redirect(url_for('bancos'))
        except Exception as e:
            return jsonify({'error': str(e)}), 500                
       

#OBTENER LOS BANCOS
@app.route('/api/obtener_bancos/', methods=['GET'])
def obtener_bancos():
    try:
        bancos = ModelBanco.get_all_bancos_no_block(db)
        
        # Crear una lista para almacenar los bancos
        bancos_data = []

        # Iterar sobre los bancos y agregarlos a la lista
        for b in bancos:
            bancos_data.append({
                'id': b.id,
                'nombre': b.nombre
            })
        
        if bancos_data:
            return jsonify(bancos_data), 200
        else:
            return jsonify({'error': 'No hay bancos disponibles'}), 404
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500




# OBTENER CFDI
@app.route('/api/obtener_cfdi/', methods=['GET'])
def obtener_cfd():

    try:
        # Obtener todos los registros de uso de CFDI desde la base de datos
        cfdi = ModelCFDI.get_all_usoCFDI(db)

        # Crear una lista para almacenar los datos de CFDI
        cfdi_data = []

        # Iterar sobre los registros de CFDI y agregarlos a la lista
        for c in cfdi:

            cfdi_data.append(
                {
                'id': c.id,
                'nombre': c.clave,
                'decripcion': c.decripcion
            })


        if cfdi_data:
            return jsonify(cfdi_data), 200
        
        else:

            return jsonify({'error': 'No hay registros de CFDI disponibles'}), 404
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def convertir_a_float(valor, valor_default=0.0):
    """
    Convierte un valor a float. Si no es posible, devuelve un valor predeterminado.
    
    :param valor: Valor a convertir
    :param valor_default: Valor por defecto si la conversión falla
    :return: Valor convertido a float o el valor por defecto
    """
    try:
        # Elimina el símbolo '$' y las comas, luego convierte a float
        return float(valor.replace("$", "").replace(",", ""))
    except (ValueError, AttributeError):
        return valor_default



# EDIT PRESUPUESTO
@app.route('/edit_presupuesto/', methods=['GET', 'POST'])
def edit_presupuesto():
    id = request.args.get('id')
    try:
        if request.method == 'POST':
            # Verifica si el token en la sesión coincide con el del formulario
            token = request.form.get('token')
            if token and session.get('token') == token:
                # Elimina el token de la sesión para evitar reenvíos duplicados
                session.pop('token', None)
         
                
                try:
                    #OBTENER ESTATUS DE PRESUPUESTO Y ESTADO DE CONTRATOS
                    estatus_contratos = request.form.get('status_contratos')
                    estatus_presupuesto = request.form.get('status_presupuesto')
                    
                    # Validar si ambos son "true"
                    if estatus_contratos == "true" and estatus_presupuesto == "true":
                        estatus_actual = 1
                    else:
                        estatus_actual = 0
                
                    
                    
                    
                    # 1- ACTUALIZAR CABECERA DE PRESUPUESTO
                    presupuesto = Presupuesto(
                        id=id,
                        proyecto=request.form['PROYECTO'],
                        presupuesto_cliente=request.form['PRESUPUESTO_CLIENTE'],
                        pagado_cliente=request.form['PAGADO_CLIENTE'],
                        porcetaje_gastado_real=request.form['PORCENTAJE_GASTADO_REAL'],
                        subtotal_cliente=request.form['subtotalCliente'],
                        subtotal_proveedor=request.form['subtotalContratista'],
                        subtotal_diferencia=request.form['subtotalDiferencia'],
                        porcentaje_indirecto=request.form.get('porcentajeIndirecto', 0),
                        total_porcentaje_indirecto=request.form.get('totalIndirectoCliente', 0),
                        total_cliente=request.form.get('totalCliente', 0),
                        total_proveedor=request.form.get('totalContratista', 0),
                        total_diferencia=request.form.get('totalDiferencia', 0),
                        usuario_id=current_user.id,
                        estatus=estatus_actual,
                        estatus_contratos = estatus_contratos,
                        estatus_presupuesto = estatus_presupuesto
                    )
                    
                   
                    
                    actualizacion_presupuesto = ModelPresupuesto.actualizar_presupuesto(db, presupuesto)
                    
                    ### ACTUALIZAR DETALLE DE PRESUPUESTO
                    
                    ### Actualizar partidas cargadas por lider de obra
                    id_detalle = request.form.getlist('ID_DETALLE[]')
                    proveedor_db = request.form.getlist('PROVEEDOR_DB[]')
                    presupuesto_cliente_db = request.form.getlist('PRESUPUESTO_CLIENTE_DB[]')
                    presupuesto_proveedor_db = request.form.getlist('PRESUPUESTO_PROVEEDOR_DB[]')
                    diferencia_presupuestos_db = request.form.getlist('DIFERENCIA_DB[]')  
                    contrato_firmado_partidas_db = request.form.getlist('CONTRATO_FIRMADO_DB[]')   
                    estatus_db = request.form.getlist('STATUS_DB[]')
                    
                    #ACTUALIZAR LAS PARTIDAD CARGADAS
                    for i in range(len(id_detalle)):
                        fila = {
                            'id_detalle': id_detalle[i],
                            'id_proveedor': proveedor_db[i],
                            'presupuesto_cliente': convertir_a_float(presupuesto_cliente_db[i]),
                            'presupuesto_contratista': convertir_a_float(presupuesto_proveedor_db[i]),
                            'diferencia': convertir_a_float(diferencia_presupuestos_db[i]),
                            'contrato_firmado': contrato_firmado_partidas_db[i],
                            'estatus': estatus_db[i],
                        }
                        
                        resultado = ModelPresupuesto.actualizar_detalle_presupuesto(db, fila)
                    
                    #ACTUALIZAR PRESUPUESTO BAUART
                    id_presupuestos_buart = request.form.getlist('ID_DETALLE_BUART_DB[]')
                    presupuesto_cliente_buart = request.form.getlist('INPUT_PRESUPUESTO_CLIENTE_BAUART_DB[]')
                    presupuesto_proveedor_buart = request.form.getlist('INPUT_PRESUPUESTO_CONTRATISTA_BAUART_DB[]')
                    diferencia_presupuestos_buart = request.form.getlist('INPUT_DIFERENCIA_PRESUPUESTOS_BUART_DB[]')
                    
                    for i in range(len(id_presupuestos_buart)):
                        fila = {
                            'id_detalle': id_presupuestos_buart[i],
                            'presupuesto_cliente': convertir_a_float(presupuesto_cliente_buart[i]),
                            'presupuesto_contratista': convertir_a_float(presupuesto_proveedor_buart[i]),
                            'diferencia': convertir_a_float(diferencia_presupuestos_buart[i]),
                        }
                        
                        resultado_actualizacion = ModelPresupuesto.actualizar_presupuesto_bauart(db, fila)
                    

                     
                    # ACTUALIZAR PARTIDAS PRESUPUESTO BUART
                    id_partidas_buart_db = request.form.getlist('ID_DETALLE_SUB_BUART_DB[]')
                    especialidades_buart_db = request.form.getlist('ESPECIALIDADES_SUB_DB[]')
                    id_concepto_partidas_buart_db = request.form.getlist('INPUT_ID_ESPECIALIDADES_SUB_DB[]')
                    id_proveedor_partidas_buart_db = request.form.getlist('PROVEEDOR_SUB_DB[]')
                    presupuesto_cliente_buart_db = request.form.getlist('SUB_PRESUPUESTO_CLIENTE_DB[]')
                    presupuesto_proveedor_buart_db = request.form.getlist('SUB_PRESUPUESTO_PROVEEDOR_DB[]')
                    diferencia_presupuestos_buart_db = request.form.getlist('SUB_DIFERENCIA_DB[]')                     
                    estatus_presupues_bauart_db = request.form.getlist('STATUS_SUB_DB[]')
                    
                    
                    
                    for i in range(len(id_partidas_buart_db)):
                        fila = {
                            'id_detalle': id_partidas_buart_db[i],
                            'especialidad': especialidades_buart_db[i],
                            'id_concepto': id_concepto_partidas_buart_db[i],
                            'id_proveedor': id_proveedor_partidas_buart_db[i],
                            'presupuesto_cliente': convertir_a_float(presupuesto_cliente_buart_db[i]),
                            'presupuesto_proveedor': convertir_a_float(presupuesto_proveedor_buart_db[i]),
                            'diferencia': convertir_a_float(diferencia_presupuestos_buart_db[i]),
                            'estatus': estatus_presupues_bauart_db[i]
                        }
                        
                        resultado_actualizacion = ModelPresupuesto.actualizar_presupuesto_bauart_detalle(db, fila)
                    
                    
                    ### GUARDAR NUEVAR PARTIDAD PRESUPUESTO BUART
                    id_partida_detalle_presupuesto = request.form.getlist('ID_PARTIDA_DB[]')
                    
                    for i in range(len(id_partida_detalle_presupuesto)):
                        # OBTENER LOS DATOS DE LA PARTIDA
                        id_concepto_nueva_partida = request.form.getlist(f'INPUT_ID_ESPECIALIDADES_SUB_{i}[]')
                        is_nomina_nueva_partida = request.form.getlist(f'IS_NOMINA_{i}[]')
                        conceptos_nueva_partida = request.form.getlist(f'ESPECIALIDADES_SUB-{i}[]')
                        proveedor_nueva_partida = request.form.getlist(f'PROVEEDOR_SUB-{i}[]')
                        presupuesto_cliente_nueva = request.form.getlist(f'SUB_PRESUPUESTO_CLIENTE-{i}[]')
                        presupuesto_proveedor_nueva = request.form.getlist(f'SUB_PRESUPUESTO_PROVEEDOR-{i}[]')
                        diferencia_presupuesto_nueva = request.form.getlist(f'SUB_DIFERENCIA-{i}[]')
                        status_nueva = request.form.getlist(f'SUB_STATUS-{i}[]')
                        
                        
                        for j in range(len(id_concepto_nueva_partida)):
                            
                            nueva_detalle = DetalleBauart(id=0,
                                    id_presupuesto_bauart=id_presupuestos_buart[j],
                                    id_concepto=id_concepto_nueva_partida[j],
                                    concepto=conceptos_nueva_partida[j],
                                    id_proveedor=proveedor_nueva_partida[j],
                                    presupuesto_cliente=convertir_a_float(presupuesto_cliente_nueva[j]),
                                    presupuesto_contratista=convertir_a_float(presupuesto_proveedor_nueva[j]),
                                    diferencia=convertir_a_float(diferencia_presupuesto_nueva[j]),
                                    is_nomina=is_nomina_nueva_partida[j],
                                    estatus=status_nueva[j]
                                    )
                            
                            # Guarda el detalle Bauart
                            ModelPresupuesto.agregar_detalle_bauart(db, nueva_detalle)
                            
                
                    #### GUARDAR NUEVAS PARTIDAS PRESUPUESTO
                    id_partida = request.form.getlist('ID_PARTIDAS_NV[]')
                    conceptos_partidas = request.form.getlist('ESPECIALIDADES_NV[]')
                    id_concepto_partidas = request.form.getlist('ID_ESPECIALIDADES_NV[]')
                    proveedor_partidas = request.form.getlist('PROVEEDOR_NV[]')
                    presupuesto_cliente_partidas = request.form.getlist('PRESUPUESTO_CLIENTE_NV[]')
                    presupuesto_proveedor_partidas = request.form.getlist('PRESUPUESTO_PROVEEDOR_NV[]')
                    diferencia_presupuestos_partidas = request.form.getlist('DIFERENCIA_NV[]')
                    contrato_firmado_partidas = request.form.getlist('CONTRATO_FIRMADO_NV[]')
                    status_partidas = request.form.getlist('STATUS_NV[]')

                    for i in range(len(id_partida)):
                        detalle = DetallePresupuesto(id=0,
                            id_presupuesto=id,
                            id_concepto=id_concepto_partidas[i],
                            concepto=conceptos_partidas[i],
                            id_proveedor=proveedor_partidas[i],
                            presupuesto_cliente=convertir_a_float(presupuesto_cliente_partidas[i]),
                            presupuesto_contratista=convertir_a_float(presupuesto_proveedor_partidas[i]),
                            diferencia=convertir_a_float(diferencia_presupuestos_partidas[i]),
                            contrato_firmado=contrato_firmado_partidas[i],
                            estatus = status_partidas[i]
                            
                        )

                        # Guarda el detalle del presupuesto
                        detalle_id = ModelPresupuesto.agregar_detalle_presupuesto(db, detalle)
                        detalle.id = detalle_id

                        # Si el proveedor es "Bauart", se genera un presupuesto Bauart
                        if detalle.id_proveedor == '0':
                            nombre_sub_presupuesto = request.form[f'NOMBRE_SUB-{id_partida[i]}']
                            total_cliente_bauart = request.form[f'INPUT_PRESUPUESTO_CLIENTE_BAUART-{id_partida[i]}']
                            total_proveedor_bauart = request.form[f'INPUT_PRESUPUESTO_CONTRATISTA_BAUART-{id_partida[i]}']
                            diferencia_bauart = request.form[f'INPUT_DIFERENCIA_PRESUPUESTOS-{id_partida[i]}']

                            presupuesto_bauart = PresupuestoBauart(
                                id=0,
                                concepto=detalle.concepto,
                                id_detalle=detalle.id,
                                nombre_presupuesto=nombre_sub_presupuesto,
                                total_presupuesto_cliente=convertir_a_float(total_cliente_bauart),
                                total_presupuesto_proveedor=convertir_a_float(total_proveedor_bauart),
                                diferencia_presupuesto=convertir_a_float(diferencia_bauart)
                            )

                            # Guarda el presupuesto Bauart
                            presupuesto_bauart_id = ModelPresupuesto.agregar_presupuesto_bauart(db, presupuesto_bauart)
                            presupuesto_bauart.id = presupuesto_bauart_id

                            # Detalles del presupuesto Bauart
                            id_concepto_sub = request.form.getlist(f'INPUT_ID_ESPECIALIDADES_SUB_{id_partida[i]}[]')
                            is_nomina_sub = request.form.getlist(f'IS_NOMINA_{id_partida[i]}[]')
                            conceptos_sub = request.form.getlist(f'ESPECIALIDADES_SUB-{id_partida[i]}[]')
                            proveedores_sub = request.form.getlist(f'PROVEEDOR_SUB-{id_partida[i]}[]')
                            sub_presupuesto_cliente = request.form.getlist(f'SUB_PRESUPUESTO_CLIENTE-{id_partida[i]}[]')
                            sub_presupuesto_proveedor = request.form.getlist(f'SUB_PRESUPUESTO_PROVEEDOR-{id_partida[i]}[]')
                            sub_diferencia = request.form.getlist(f'SUB_DIFERENCIA-{id_partida[i]}[]')

                            for j in range(len(conceptos_sub)):
                                detalle_bauart = DetalleBauart(id=0,
                                    id_presupuesto_bauart=presupuesto_bauart.id,
                                    id_concepto=id_concepto_sub[j],
                                    concepto=conceptos_sub[j],
                                    id_proveedor=proveedores_sub[j],
                                    presupuesto_cliente=convertir_a_float(sub_presupuesto_cliente[j]),
                                    presupuesto_contratista=convertir_a_float(sub_presupuesto_proveedor[j]),
                                    diferencia=convertir_a_float(sub_diferencia[j]),
                                    is_nomina=is_nomina_sub[j]
                                )

                                # Guarda el detalle Bauart
                                ModelPresupuesto.agregar_detalle_bauart(db, detalle_bauart)

                    return redirect(url_for('prespuestos'))

                            
              
                
                    
                
                except BadRequestKeyError as e:
                    return jsonify({"error": f"Clave faltante en el formulario: {e}"}), 400
                
                except Exception as e:
                    return jsonify({"error": f"Ocurrió un error al actualizar el presupuesto: {e}"}), 500
                
            else:
                return redirect(url_for('prespuestos'))
        else:
            session['token'] = str(uuid.uuid4())
            proveedores = ModelProveedores.get_all_proveedores_not_blocked(db)
            puestos = ModelPuesto.get_all_puestos_no_block(db)
            presupuesto = ModelPresupuesto.obtener_presupuesto_completo(db, id)
            return render_template('modificar_presupuesto.html', presupuesto=presupuesto, proveedores=proveedores, puestos=puestos,token=session['token'])
    
    except Exception as e:
        db.rollback()
        return jsonify({"error": f"Ocurrió un error en edit_presupuesto: {e}"}), 500


# PRESUPUESTO
@app.route('/Presupuestos', methods=['GET', 'POST'] )
def prespuestos():
    
    if request.method == 'POST':
        pass
    
    else:
        presupuestos = ModelPresupuesto.obtener_presupuestos(db)
        
        return render_template('presupuestos.html',presupuestos=presupuestos)


# ALTA PRESUPUESTOS
@app.route('/altaPresupuestos', methods=['GET', 'POST'])
def altaPresupuestos():
    if request.method == 'POST':
        # Verifica si el token en la sesión coincide con el del formulario
        token = request.form.get('token')
        if token and session.get('token') == token:
            # Elimina el token de la sesión para evitar reenvíos duplicados
            session.pop('token', None)
        

            # CABECERA DE PRESUPUESTO
            nuevo_presupuesto = Presupuesto(
                id=0,
                proyecto=request.form['PROYECTO'],
                id_proyecto = int(request.form['PROYECTO_ID']),
                id_director=int(request.form['ID_DIRECTOR']),
                id_cliente=int(request.form['CLIENTE_ID']),
                id_empresa=current_user.id_empresa,
                presupuesto_cliente=convertir_a_float(request.form['PRESUPUESTO_CLIENTE']),
                falta_por_cobrar=convertir_a_float(request.form['FALTA_POR_COBRAR']),
                porcentaje_por_cobra=convertir_a_float(request.form['PORCENTAJE_POR_COBRAR']),
                fecha_inicio=request.form['FECHA_INICIO'],
                direccion_obra=request.form['DIRECCION'],
                pagado_cliente=convertir_a_float(request.form['PAGADO_CLIENTE']),
                porcentaje_pagado_cliente=float(request.form['PORCENTAJE_PAGADO_CLIENTE']),
                falta_por_gastar=convertir_a_float(request.form['FALTA_POR_GASTAR']),
                porcentaje_por_gastar=float(request.form['PORCENTAJE_POR_GASTAR']),
                fecha_fin=request.form['FECHA_FIN'],
                director_obra=request.form['DIRECTOR'],
                gastado_real=convertir_a_float(request.form['GASTADO_REAL']),
                porcetaje_gastado_real=float(request.form['PORCENTAJE_GASTADO_REAL']),
                estatus_proyecto=convertir_a_float(request.form['ESTATUS_OBRA']),
                total_semanas=int(request.form['DIAS_TOTALES']),
                subtotal_cliente=convertir_a_float(request.form['subtotalCliente']),
                subtotal_proveedor=convertir_a_float(request.form['subtotalContratista']),
                subtotal_diferencia=convertir_a_float(request.form['subtotalDiferencia']),
                porcentaje_indirecto=0,
                total_porcentaje_indirecto=0,
                total_cliente=0,
                total_proveedor=0,
                total_diferencia=0,
                usuario_id=current_user.id,
                estatus=0
            )
                        
            

            # Guarda el presupuesto en la base de datos
            presupuesto_id = ModelPresupuesto.crear_presupuesto(db, nuevo_presupuesto)
            nuevo_presupuesto.id = presupuesto_id

            # PARTIDAS PRESUPUESTALES
            id_partida = request.form.getlist('ID_PARTIDAS[]')
            conceptos_partidas = request.form.getlist('ESPECIALIDADES[]')
            id_concepto_partidas = request.form.getlist('ID_ESPECIALIDADES[]')
            proveedor_partidas = request.form.getlist('PROVEEDOR[]')
            presupuesto_cliente_partidas = request.form.getlist('PRESUPUESTO_CLIENTE[]')
            presupuesto_proveedor_partidas = request.form.getlist('PRESUPUESTO_PROVEEDOR[]')
            diferencia_presupuestos_partidas = request.form.getlist('DIFERENCIA[]')
            contrato_firmado_partidas = request.form.getlist('CONTRATO_FIRMADO[]')

            for i in range(len(id_partida)):
                detalle = DetallePresupuesto(id=0,
                    id_presupuesto=nuevo_presupuesto.id,
                    id_concepto=id_concepto_partidas[i],
                    concepto=conceptos_partidas[i],
                    id_proveedor=proveedor_partidas[i],
                    presupuesto_cliente=convertir_a_float(presupuesto_cliente_partidas[i]),
                    presupuesto_contratista=convertir_a_float(presupuesto_proveedor_partidas[i]),
                    diferencia=convertir_a_float(diferencia_presupuestos_partidas[i]),
                    contrato_firmado=contrato_firmado_partidas[i] == '1'
                )

                # Guarda el detalle del presupuesto
                detalle_id = ModelPresupuesto.agregar_detalle_presupuesto(db, detalle)
                detalle.id = detalle_id

                # Si el proveedor es "Bauart", se genera un presupuesto Bauart
                if detalle.id_proveedor == '0':
                    nombre_sub_presupuesto = request.form[f'NOMBRE_SUB-{id_partida[i]}']
                    total_cliente_bauart = request.form[f'INPUT_PRESUPUESTO_CLIENTE_BAUART-{id_partida[i]}']
                    total_proveedor_bauart = request.form[f'INPUT_PRESUPUESTO_CONTRATISTA_BAUART-{id_partida[i]}']
                    diferencia_bauart = request.form[f'INPUT_DIFERENCIA_PRESUPUESTOS-{id_partida[i]}']

                    presupuesto_bauart = PresupuestoBauart(
                        id=0,
                        concepto=detalle.concepto,
                        id_detalle=detalle.id,
                        nombre_presupuesto=nombre_sub_presupuesto,
                        total_presupuesto_cliente=convertir_a_float(total_cliente_bauart),
                        total_presupuesto_proveedor=convertir_a_float(total_proveedor_bauart),
                        diferencia_presupuesto=convertir_a_float(diferencia_bauart)
                    )

                    # Guarda el presupuesto Bauart
                    presupuesto_bauart_id = ModelPresupuesto.agregar_presupuesto_bauart(db, presupuesto_bauart)
                    presupuesto_bauart.id = presupuesto_bauart_id

                    # Detalles del presupuesto Bauart
                    id_concepto_sub = request.form.getlist(f'INPUT_ID_ESPECIALIDADES_SUB_{id_partida[i]}[]')
                    is_nomina_sub = request.form.getlist(f'IS_NOMINA_{id_partida[i]}[]')
                    conceptos_sub = request.form.getlist(f'ESPECIALIDADES_SUB-{id_partida[i]}[]')
                    proveedores_sub = request.form.getlist(f'PROVEEDOR_SUB-{id_partida[i]}[]')
                    sub_presupuesto_cliente = request.form.getlist(f'SUB_PRESUPUESTO_CLIENTE-{id_partida[i]}[]')
                    sub_presupuesto_proveedor = request.form.getlist(f'SUB_PRESUPUESTO_PROVEEDOR-{id_partida[i]}[]')
                    sub_diferencia = request.form.getlist(f'SUB_DIFERENCIA-{id_partida[i]}[]')

                    for j in range(len(conceptos_sub)):
                        detalle_bauart = DetalleBauart(id=0,
                            id_presupuesto_bauart=presupuesto_bauart.id,
                            id_concepto=id_concepto_sub[j],
                            concepto=conceptos_sub[j],
                            id_proveedor=proveedores_sub[j],
                            presupuesto_cliente=convertir_a_float(sub_presupuesto_cliente[j]),
                            presupuesto_contratista=convertir_a_float(sub_presupuesto_proveedor[j]),
                            diferencia=convertir_a_float(sub_diferencia[j]),
                            is_nomina=is_nomina_sub[j]
                        )

                        # Guarda el detalle Bauart
                        ModelPresupuesto.agregar_detalle_bauart(db, detalle_bauart)

            return redirect(url_for('prespuestos'))

    else:
        session['token'] = str(uuid.uuid4())
        directores = ModelEmpleado.obtener_directores(db)
        return render_template('altaPresupuestos.html', token=session['token'],directores=directores)


# PAGOS SUGERIDOS
@app.route('/carga_pagos_sugeridos/', methods=['GET', 'POST'])
def carga_pagos_sugeridos():
    id = request.args.get('id')
    if request.method == 'POST':
        pass
    
    else:
        session['token'] = str(uuid.uuid4())
        proveedores = ModelProveedores.get_all_proveedores_not_blocked(db)
        puestos = ModelPuesto.get_all_puestos_no_block(db)
        presupuesto = ModelPresupuesto.obtener_presupuesto_completo(db, id)
        return render_template('altaPagosSugeridos.html', presupuesto=presupuesto, proveedores=proveedores, puestos=puestos,token=session['token'])


@app.route('/pagos_sugeridos', methods=['GET', 'POST'])
def pagos_sugeridos():
   
    if request.method == 'POST':
        pass
    
    else:
        print("hola mundo")
        presupuestos = ModelPresupuesto.obtener_presupuestos(db)    
        return render_template('pagos_sugeridos.html',presupuestos=presupuestos)


# MATERIALES
@app.route('/materiales', methods=['GET', 'POST'])
def materiales():
    if request.method == 'POST':
        # Validación de los campos requeridos
        familia_id = request.form.get('FamiliaID')
        descripcion_material = request.form.get('descripcionMaterial')
        medidaMaterial = request.form.get('medidaMaterial')


        if not familia_id or not descripcion_material:
            # Puedes manejar el error de una manera apropiada
            flash("Por favor, complete todos los campos requeridos.", "error")
            return redirect(url_for('materiales'))

        new_material = MaterialesFamilia(
            # No es necesario establecer el id si es auto-incremental
            id=0,
            id_familia=familia_id,
            material=descripcion_material,
            unidad_medida=medidaMaterial
        )

        try:
            ModelMaterialesFamilia.crear_material_familia(db, new_material)
            flash("Material creado con éxito.", "success")  # Mensaje de éxito
        except Exception as e:
            print(f"Error al crear material: {str(e)}")  # Impresión para depuración
            flash("Error al crear material. Inténtelo de nuevo.", "error")  # Mensaje de error
            return redirect(url_for('materiales'))

        # Redireccionar después de procesar el formulario
        return redirect(url_for('materiales'))


    materiales = ModelMaterialesFamilia.get_all_materiales_familia(db)
    familias = ModelFamilias.get_all_familias_not_block(db)
    return render_template('materiales.html', materiales=materiales, familias=familias)


# FILTRAR MATERIALES
@app.route('/filtro_materiales', methods=['POST'])
@login_required
def filtro_materiales():
    especialidad = request.form.get('especialidad', '')
    estatus = request.form.get('estatus', '')

    try:
        especialidad = str(especialidad)
        especialidades = ModelEspecialidades.filter_especialidad(db,especialidad, estatus)
       
        if especialidades:
            return render_template('especialidades.html', especialidades=especialidades)
        else:
            return redirect(url_for('especialidades'))
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500



# EDITAR MATERIALES
@app.route('/edit_materiales/<int:id>', methods=['POST'])
def edit_materiales(id):
       
    familia_id = request.form.get('inputIdFamiliaEdit')
    familia = request.form.get('inputFamiliaEdit')
    descripcion_material = request.form.get('inputDecripcion')
    medidaMaterial = request.form.get('inputUOMEdit')
        
    if not familia_id or not descripcion_material:
        # Puedes manejar el error de una manera apropiada
        flash("Por favor, complete todos los campos requeridos.", "error")
        return redirect(url_for('materiales'))

    material = MaterialesFamilia(
        # No es necesario establecer el id si es auto-incremental
        id=id,
        id_familia=familia_id,
        material=descripcion_material,
        familia=familia,
        unidad_medida=medidaMaterial,
        usuario=current_user.id
    )
        
    ModelMaterialesFamilia.update_material_familia(db,material)
        
    return redirect(url_for('materiales'))


# EDITAR FAMILIA
@app.route('/edit_familia/<int:id>', methods=['POST'])
def edit_familia(id):
       

    familia = request.form.get('inputFamiliaEdit')
 

    if not familia:
        # Puedes manejar el error de una manera apropiada
        flash("Por favor, complete todos los campos requeridos.", "error")
        return redirect(url_for('familias'))

    material = Familias(
        id=id,
        familia=familia,
        fecha_registro= datetime.today(),
        usuario=current_user.id
    )
        
    ModelFamilias.update_familia(db,material)
        
    return redirect(url_for('familias'))
        
# FAMILIAS
@app.route('/familias', methods=['GET', 'POST'])
def familias():
    if request.method == 'POST':
        new_familia = Familias(
            id=0,
            familia=request.form['NuevaFamilia'],
        )
        ModelFamilias.crearFamilia(db, new_familia)
        
        # Redireccionar después de procesar el formulario para evitar duplicaciones en la actualización
        return redirect(url_for('familias'))

    familias = ModelFamilias.get_all_familias(db)

    # Imprimir los valores de is_blocked para depuración
    for familia in familias:
        print(f"Familia: {familia.familia}, is_blocked: {familia.is_blocked}")

    return render_template('familias.html', familias=familias)



#BLOQUEAR FAMILIA
@app.route('/block_familia/<int:id>', methods=['POST'])
def block_familia(id):
    if request.method == 'POST':
        try:
       
            ModelFamilias.change_status(db, id, True)
            return redirect(url_for('familias'))
        except Exception as e:
            return jsonify({'error': str(e)}), 500

#DESBLOQUEA FAMILIA
@app.route('/unblock_familia/<int:id>', methods=['POST'])
def unblock_familia(id):
    if request.method == 'POST':
        try:
          
            ModelFamilias.change_status(db, id, False)
            return redirect(url_for('familias'))
        except Exception as e:
            return jsonify({'error': str(e)}), 500        

#BLOQUEAR MATERIAL
@app.route('/block_material/<int:id>', methods=['POST'])
def block_material(id):
    if request.method == 'POST':
        try:
       
            ModelMaterialesFamilia.change_status(db, id, True)
            return redirect(url_for('materiales'))
        except Exception as e:
            return jsonify({'error': str(e)}), 500

#DESBLOQUEA MATERIAL
@app.route('/unblock_material/<int:id>', methods=['POST'])
def unblock_material(id):
    if request.method == 'POST':
        try:
          
            ModelMaterialesFamilia.change_status(db, id, False)
            return redirect(url_for('materiales'))
        except Exception as e:
            return jsonify({'error': str(e)}), 500                


#Orden de compra
@app.route('/api/buscar_proveedor', methods=['GET'])
def buscar_proveedor():
    query = request.args.get('query', '').strip()
    if not query:
        return jsonify([])

    try:
        connection = db_sql_server.get_connection()
        cursor = connection.cursor()

        sql_query = """
            SELECT TOP 10 ID, RAZON_SOCIAL, RFC
            FROM PROVEEDORES
            WHERE RAZON_SOCIAL LIKE ? OR RFC LIKE ?
            ORDER BY RAZON_SOCIAL
        """
        cursor.execute(sql_query, (f'%{query}%', f'%{query}%'))

        proveedores = [
            {'id': row.ID, 'razon_social': row.RAZON_SOCIAL, 'rfc': row.RFC}
            for row in cursor.fetchall()
        ]

        cursor.close()
        connection.close()

        return jsonify(proveedores)

    except Exception as e:
        print(f"Error al buscar proveedores: {e}")
        return jsonify({'error': 'Error al buscar proveedores'}), 500

ordenes = {}

@app.route('/update_requisicion_date/<int:id>', methods=['POST'])
def update_requisicion_date(id):
    """
    Actualiza el estado de una requisición y su fecha de llegada si se marca como recibida.
    Si el estado es "Pendiente", establece la fecha de llegada como NULL.
    """
    db = db_sql_server.get_connection()
    if db is None:
        return jsonify({"error": "Database connection failed"}), 500

    try:
        # Obtener el nuevo estado desde el formulario
        new_status = request.form.get('status')
        if new_status is None:
            return jsonify({"error": "Estado no especificado"}), 400

        new_status = int(new_status)

        if new_status == 1:  # Estado "Recibido"
            fecha_llegada = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            query = "UPDATE REQUISICIONES SET STATUS = ?, FECHA_LLEGADA = ? WHERE ID_REQUISICION = ?"
            with db.cursor() as cursor:
                cursor.execute(query, (new_status, fecha_llegada, id))
                db.commit()
        else:  # Estado "Pendiente", establecer FECHA_LLEGADA como NULL
            query = "UPDATE REQUISICIONES SET STATUS = ?, FECHA_LLEGADA = NULL WHERE ID_REQUISICION = ?"
            with db.cursor() as cursor:
                cursor.execute(query, (new_status, id))
                db.commit()

        return jsonify({"success": True, "message": "Requisición actualizada con éxito."}), 200
    except Exception as e:
        print(f"Error al actualizar la requisición: {e}")
        return jsonify({"error": "Error al actualizar la requisición"}), 500


@app.route('/api/ordenes_requisicion', methods=['GET'])
def get_ordenes_requisicion():
    id_requisicion = request.args.get('id_requisicion', type=int)

    if not id_requisicion:
        return jsonify({"error": "El parámetro 'id_requisicion' es obligatorio."}), 400

    db = db = db_sql_server.get_connection()
    try:
        # Consulta para obtener las órdenes relacionadas con la requisición y el nombre del proveedor
        cursor = db.cursor()
        query = """
            SELECT 
                o.ID AS id_orden,
                o.ID_REQUISICION AS id_requisicion,
                o.ID_PROVEEDOR AS id_proveedor,
                p.RAZON_SOCIAL AS proveedor_nombre  -- Nombre del proveedor
            FROM ORDENES o
            JOIN PROVEEDORES p ON o.ID_PROVEEDOR = p.ID
            WHERE o.ID_REQUISICION = ?
        """
        cursor.execute(query, (id_requisicion,))
        results = cursor.fetchall()

        # Estructurar el resultado en JSON
        ordenes = []
        for row in results:
            orden = {
                "id": row.id_orden,
                "id_requisicion": row.id_requisicion,
                "id_proveedor": row.id_proveedor,
                "proveedor_nombre": row.proveedor_nombre  # Incluir el nombre del proveedor
            }
            ordenes.append(orden)

        return jsonify(ordenes), 200
    except Exception as e:
        print(f"Error al obtener órdenes: {e}")
        return jsonify({"error": "Error al obtener las órdenes."}), 500
    finally:
        db.close()

@app.route('/guardar_orden', methods=['POST'])
def guardar_orden():
    try:
        data = request.json
        
        # Extraer datos del JSON
        id_requisicion = data['id_requisicion']
        id_proveedor = data['id_proveedor']
        fecha_hora_entrega = data.get('fecha_hora_entrega')  # Fecha y hora de entrega
        direccion_entrega = data.get('direccion_entrega')  # Dirección de entrega
        contacto = data.get('contacto')  # Contacto de la entrega
        telefono = data.get('telefono')  # Teléfono del contacto
        porcentaje_descuento = data.get('porcentaje_descuento', 0)  # Descuento, default 0 si no está presente
        numero_cotizacion = data.get('numero_cotizacion')  # Nuevo campo: Número de cotización

        # Validar que los datos necesarios estén presentes
        if not id_requisicion or not id_proveedor or not numero_cotizacion:
            return jsonify({"success": False, "message": "Datos obligatorios faltantes (requisición, proveedor o número de cotización)"}), 400

        # Validar y formatear la fecha/hora de entrega
        if fecha_hora_entrega:
            try:
                fecha_hora_entrega = datetime.strptime(fecha_hora_entrega, '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S')
            except Exception:
                return jsonify({"success": False, "message": "Formato de fecha/hora inválido"}), 400

        # Obtener conexión a la base de datos
        conn = db_sql_server.get_connection()

        # Usar el modelo para guardar la orden
        id_orden = MyOrdendeCompra.guardar_orden(
            conn,
            id_requisicion=id_requisicion,
            id_proveedor=id_proveedor,
            fecha_hora_entrega=fecha_hora_entrega,
            direccion_entrega=direccion_entrega,
            contacto=contacto,
            telefono=telefono,
            porcentaje_descuento=porcentaje_descuento,
            numero_cotizacion=numero_cotizacion  # Pasar el nuevo campo
        )

        return jsonify({"success": True, "id_orden": id_orden})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

    finally:
        if 'conn' in locals():
            conn.close()


@app.route('/actualizar_partidas', methods=['POST'])
def actualizar_partidas():
    try:
        data = request.json
        if not data:
            print("Error: No se recibió ningún dato en el request.")
            return jsonify({"success": False, "message": "No se recibió ningún dato"}), 400

        id_orden = data.get('id_orden')
        partidas = data.get('partidas')  # Lista de objetos con id y precio_unitario

        # Validaciones básicas
        if not isinstance(id_orden, int):
            print(f"Error: ID de Orden inválido. Recibido: {id_orden}")
            return jsonify({"success": False, "message": "ID de Orden inválido"}), 400
        if not isinstance(partidas, list) or not all(isinstance(p, dict) for p in partidas):
            print(f"Error: Formato de partidas inválido. Recibido: {partidas}")
            return jsonify({"success": False, "message": "Formato de partidas inválido"}), 400

        # Conexión a la base de datos
        conn = db_sql_server.get_connection()
        print(f"Conexión a la base de datos establecida. ID de Orden: {id_orden}")

        # Llamar al modelo para actualizar las partidas
        resultado = MyOrdendeCompra.actualizar_partidas(conn, id_orden, partidas)

        if resultado:
            print(f"Partidas actualizadas correctamente para la Orden: {id_orden}")
            return jsonify({"success": True})
        else:
            print(f"Error desconocido al intentar actualizar las partidas para la Orden: {id_orden}")
            return jsonify({"success": False, "message": "Error al actualizar partidas"}), 500

    except Exception as e:
        print(f"Error al actualizar partidas: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

    finally:
        if 'conn' in locals():
            conn.close()
            print("Conexión a la base de datos cerrada.")


@app.route('/generar_pdf/<int:id>', methods=['GET'])
def generar_pdf(id):
    try:
        proveedor_id = request.args.get('proveedor_id', default=None, type=int)
        pdf = ModelPDF.generar_pdf(db, id, proveedor_id)
        
        response = make_response(pdf)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'inline; filename=orden_compra_{id}.pdf'
        return response

    except Exception as e:
        return jsonify({"error": "No se pudo generar el PDF", "details": str(e)}), 500

@app.route('/api/get_bauart_presupuestos/<int:proyecto_id>', methods=['GET'])
def get_bauart_presupuestos(proyecto_id):
    """
    Endpoint para obtener presupuestos Bauart asociados a un proyecto.
    """
    try:
        print(f"Procesando proyecto_id: {proyecto_id}")
        presupuestos = ModelCompras.obtener_presupuestos_bauart(db, proyecto_id)
        print(f"Presupuestos obtenidos: {presupuestos}")
        return jsonify(presupuestos), 200
    except Exception as e:
        print(f"Error al obtener presupuestos Bauart: {e}")
        return jsonify({"error": f"Error al obtener presupuestos Bauart: {e}"}), 500


@app.route('/api/get_bauart_conceptos/<int:proyecto_id>/<int:presupuesto_id>/<int:detalle_id>', defaults={'id_familia': None}, methods=['GET'])
@app.route('/api/get_bauart_conceptos/<int:proyecto_id>/<int:presupuesto_id>/<int:detalle_id>/<int:id_familia>', methods=['GET'])
def get_bauart_conceptos(proyecto_id, presupuesto_id, detalle_id, id_familia):
    try:
        print(f"[INFO] Ruta accedida: proyecto_id={proyecto_id}, presupuesto_id={presupuesto_id}, detalle_id={detalle_id}, id_familia={id_familia}")
        
        # Paso 1: Verificar si el presupuesto Bauart existe
        print(f"[INFO] Buscando presupuesto Bauart para detalle_id={detalle_id}")
        presupuesto_bauart = ModelCompras.get_bauart_by_detalle(db, detalle_id)
        if not presupuesto_bauart:
            print(f"[ERROR] No hay presupuesto Bauart para detalle_id={detalle_id}")
            return jsonify({"error": f"No hay presupuesto Bauart asociado con ID de detalle {detalle_id}"}), 404
        print(f"[SUCCESS] Presupuesto Bauart encontrado: {presupuesto_bauart}")
        
        # Paso 2: Obtener los detalles Bauart
        print(f"[INFO] Obteniendo detalles Bauart para presupuesto_id={presupuesto_id}, id_familia={id_familia}")
        detalles_bauart = ModelCompras.get_detalles_bauart_by_presupuesto(db, presupuesto_id, id_familia=id_familia)
        if not detalles_bauart:
            print(f"[ERROR] No se encontraron detalles para presupuesto_id={presupuesto_id}, id_familia={id_familia}")
            return jsonify({"error": f"No se encontraron detalles para el presupuesto Bauart con ID {presupuesto_id}"}), 404
        print(f"[SUCCESS] Detalles Bauart obtenidos: {detalles_bauart}")

        # Paso 3: Procesar los conceptos Bauart
        print(f"[INFO] Procesando detalles Bauart para construir la respuesta JSON")
        conceptos_bauart = [
            {
                "id_concepto": detalle["id_concepto"],
                "nombre": detalle["nombre_concepto"],
                "unidad_medida": detalle["unidad_medida"],
                "id_familia": detalle["id_familia"],
                "nombre_familia": detalle.get("nombre_familia", "N/A")  # Manejo de None
            }
            for detalle in detalles_bauart
        ]
        print(f"[SUCCESS] Conceptos Bauart procesados: {conceptos_bauart}")

        # Respuesta final
        return jsonify({"conceptos_bauart": conceptos_bauart}), 200

    except Exception as e:
        print(f"[CRITICAL] Error interno del servidor: {e}")
        return jsonify({"error": f"Error interno del servidor: {e}"}), 500


    
@app.route('/api/get_material_by_name/<string:material_name>', methods=['GET'])
def get_material_by_name(material_name):
    try:
        db = db_sql_server.get_connection()
        material_name = material_name.strip()  # Limpia posibles espacios en blanco
        # Usamos el método simplificado para buscar el material
        material = ModelMateriales.get_material_by_name(db, material_name)
        if not material:
            print(f"Material '{material_name}' no encontrado o bloqueado.")
            return jsonify({"error": f"Material '{material_name}' no encontrado o bloqueado."}), 404
        return jsonify(material), 200
    except Exception as e:
        print(f"Error en get_material_by_name: {e}")
        return jsonify({"error": "Error interno del servidor"}), 500




@app.route('/generar-orden/<int:id>', methods=['GET'])
def generar_orden_view(id):
    try:
        db = db_sql_server.get_connection()
        if not db:
            return jsonify({"error": "Error al conectar con la base de datos"}), 500

        # Obtén la requisición
        requisicion = MyOrdendeCompra.get_requisicion_by_id(db, id)
        if not requisicion:
            return jsonify({"error": "Requisición no encontrada"}), 404

        # Obtén las partidas de la requisición
        partidas = MyOrdendeCompra.get_partidas_by_requisicion_id(db, id)
        print(f"Requisición: {requisicion}")
        print(f"Partidas: {partidas}")

        # Renderiza la vista con los datos
        return render_template('orden_compra.html', requisicion=requisicion, partidas=partidas)

    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"error": "Error al generar la orden de compra"}), 500

    finally:
        db.close()

@app.route('/generar_orden/<int:id_requisicion>', methods=['GET'])
def generar_orden(id_requisicion):
    db = db_sql_server.get_connection()

    # Verificar si la requisición existe
    requisicion = MyOrdendeCompra.get_requisicion_by_id(db, id_requisicion)
    if not requisicion:
        return "Requisición no encontrada", 404

    # Verificar si ya existe una orden vinculada
    query = "SELECT ID FROM ORDENES WHERE ID_REQUISICION = ?"
    cursor = db.cursor()
    cursor.execute(query, (id_requisicion,))
    orden_existente = cursor.fetchone()

    if orden_existente:
        id_orden = orden_existente[0]
    else:
        # Crear una nueva orden automáticamente
        id_orden = MyOrdendeCompra.create_orden(db, id_requisicion)

    # Obtener partidas asociadas a la requisición
    partidas = MyOrdendeCompra.get_partidas_by_requisicion_id(db, id_requisicion)

    # Renderizar la página con los datos
    return render_template(
        'orden.html', 
        requisicion=requisicion, 
        partidas=partidas, 
        orden_id=id_orden
    )




def get_partidas_for_order(order_id):
    db = db_sql_server.get_connection()
    cursor = db.cursor()
    query = """
        SELECT id, descripcion, unidad, cantidad, precio
        FROM PARTIDAS_ORDEN
        WHERE id_orden = ?
    """
    cursor.execute(query, (order_id,))
    partidas = cursor.fetchall()
    db.close()
    return [{"id": p[0], "descripcion": p[1], "unidad": p[2], "cantidad": p[3], "precio": p[4]} for p in partidas]


def get_requisicion_details(order_id):
    db = db_sql_server.get_connection()
    cursor = db.cursor()
    query = """
        SELECT r.id AS id_requisicion, r.nombre_proyecto, r.concepto
        FROM REQUISICIONES r
        JOIN ORDENES o ON r.id = o.ID_REQUISICION
        WHERE o.ID = ?
    """
    cursor.execute(query, (order_id,))
    requisicion = cursor.fetchone()
    db.close()
    if requisicion:
        return {"id_requisicion": requisicion[0], "nombre_proyecto": requisicion[1], "concepto": requisicion[2]}
    return None


def generar_pdf_orden_compra(requisicion, partidas):
    filename = f"Orden_Compra_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    c = canvas.Canvas(filename, pagesize=A4)
    width, height = A4

    # Título
    c.setFont("Helvetica-Bold", 16)
    c.drawString(1 * inch, height - 1 * inch, "Orden de Compra - Detalles")

    # Detalles de la requisición
    y = height - 1.5 * inch
    c.setFont("Helvetica", 10)
    c.drawString(1 * inch, y, f"ID Requisición: {requisicion['id_requisicion']}")
    c.drawString(1 * inch, y - 0.2 * inch, f"Proyecto: {requisicion['nombre_proyecto']}")
    c.drawString(1 * inch, y - 0.4 * inch, f"Concepto: {requisicion['concepto']}")

    # Detalles de las partidas
    y -= 0.8 * inch
    c.setFont("Helvetica-Bold", 10)
    c.drawString(1 * inch, y, "Descripción")
    c.drawString(3.5 * inch, y, "Unidad")
    c.drawString(4.5 * inch, y, "Cantidad")
    c.drawString(5.5 * inch, y, "Precio Unitario")
    c.drawString(6.5 * inch, y, "Subtotal")
    y -= 0.3 * inch

    c.setFont("Helvetica", 10)
    total = 0
    for partida in partidas:
        descripcion = partida["descripcion"]
        unidad = partida["unidad"]
        cantidad = float(partida["cantidad"])
        precio = float(partida.get("precio", 0))
        subtotal = cantidad * precio
        total += subtotal

        c.drawString(1 * inch, y, descripcion)
        c.drawString(3.5 * inch, y, unidad)
        c.drawString(4.5 * inch, y, str(cantidad))
        c.drawString(5.5 * inch, y, f"${precio:.2f}")
        c.drawString(6.5 * inch, y, f"${subtotal:.2f}")
        y -= 0.3 * inch

        if y < inch:
            c.showPage()
            y = height - 1 * inch

    # Total
    c.setFont("Helvetica-Bold", 12)
    c.drawString(5.5 * inch, y - 0.3 * inch, f"Total: ${total:.2f}")

    c.save()
    return filename

#FlujoRequisiones Y Partida Requisiciones
@app.route('/api/nueva_requisicion', methods=['POST'])
def nueva_requisicion():
    # Obtener datos desde el JSON usando `request.json`
    data = request.json

    nombre_proyecto = data.get("nombre_proyecto")
    concepto = data.get("concepto")
    fecha_requerida = data.get("fecha_requerida")
    status = data.get("status")

    # Validación para los campos obligatorios
    if not nombre_proyecto or not concepto:
        return jsonify({"error": "Los campos 'nombre_proyecto' y 'concepto' son obligatorios"}), 400

    # Agregar valores predeterminados o adicionales al diccionario `data`
    data["fecha_requerida"] = fecha_requerida  # Puede ser `None` o una fecha específica
    data["status"] = status  # Puede ser `None` o un valor específico

    # Llamar a `create_requisicion` para crear la requisición
    id_requisicion = ModelRequisiciones.create_requisicion(db, data)

    # Comprobar si la creación fue exitosa y devolver una respuesta adecuada
    if id_requisicion:
        return jsonify({"id": id_requisicion, "message": "Requisición creada exitosamente"}), 201
    else:
        return jsonify({"error": "No se pudo crear la requisición"}), 500
    
@app.route('/api/agregar_partida_requisicion', methods=['POST'])
def agregar_partida_requisicion():
    data = request.json
    partidas = data.get('partidas', [])
    
    if not partidas:
        return jsonify({"error": "No se recibieron partidas"}), 400

    for partida in partidas:
        # Validar campos obligatorios en cada partida
        if not partida.get("id_requisicion") or not partida.get("descripcion") or not partida.get("unidad") or not partida.get("cantidad"):
            return jsonify({"error": "Cada partida debe tener 'id_requisicion', 'descripcion', 'unidad' y 'cantidad'"}), 400

        # Validar los nuevos campos (MONEDA y TIPO_CAMBIO)
        moneda = partida.get("moneda")
        tipo_cambio = partida.get("tipo_cambio")

        if not moneda or moneda not in ["USD", "MXN"]:
            return jsonify({"error": "Cada partida debe tener 'moneda' válida ('USD' o 'MXN')"}), 400

        if tipo_cambio is None or not isinstance(tipo_cambio, (int, float)):
            return jsonify({"error": "Cada partida debe tener un 'tipo_cambio' numérico válido"}), 400

        # Llamar a la función `add_partida` para agregar cada partida
        resultado = ModelRequisiciones.add_partida(db, partida)
        if not resultado:
            return jsonify({"error": "Ocurrió un error al agregar una o más partidas"}), 500
        
    return jsonify({"message": "Todas las partidas agregadas exitosamente"}), 201

@app.route('/ConsultarRequisiciones', methods=['GET', 'POST'])
def ConsultarRequisicionesPartidas_view():
    proyecto = request.form.get('proyecto', '').strip() if request.method == 'POST' else None
    concepto = request.form.get('concepto', '').strip() if request.method == 'POST' else None
    estatus = request.form.get('estatus', '').strip() if request.method == 'POST' else None

    status = int(estatus) if estatus and estatus.isdigit() else None

    db = db_sql_server.get_connection()
    if db is None:
        return "Error al conectar con la base de datos.", 500

    try:
        requisiciones = ModelRequisiciones.filter_requisiciones(db, nombre_proyecto=proyecto, concepto=concepto, status=status)
    finally:
        db.close()

    return render_template('menurequisiciones.html', requisiciones=requisiciones)


@app.route('/update_requisicion_status/<int:id>', methods=['POST'])
def update_requisicion_status(id):
    new_status = request.form.get('status')  # Get the new status from the form data
    if new_status is None:
        flash("Invalid status value.")
        return redirect(url_for('ConsultarRequisicionesPartidas_view'))
    
    try:
        db = db_sql_server.get_connection()
        ModelRequisiciones.update_status(db, id, int(new_status))
        db.commit()  # Commit changes at the connection level
        flash("Requisition status updated successfully.")
    except Exception as e:
        flash(f"Error updating requisition status: {str(e)}")
    finally:
        db.close()
    
    return redirect(url_for('ConsultarRequisicionesPartidas_view'))


@app.route('/delete_requisicion/<int:requisicion_id>', methods=['DELETE'])
def delete_requisicion(requisicion_id):
    db = db_sql_server.get_connection()
    if db is None:
        return jsonify({"error": "Database connection failed"}), 500

    try:
        with db.cursor() as cursor:
            # First, delete partitions associated with the requisition
            cursor.execute("DELETE FROM PARTIDAS_REQUISICION WHERE ID_REQUISICION = ?", (requisicion_id,))
            # Then, delete the requisition itself
            cursor.execute("DELETE FROM REQUISICIONES WHERE ID_REQUISICION = ?", (requisicion_id,))
            db.commit()  # Commit both deletions
        return jsonify({"success": "Requisicion and its partitions deleted successfully."}), 200
    except Exception as e:
        print(f"Error deleting requisition: {e}")
        return jsonify({"error": "Error deleting requisition"}), 500
    finally:
        db.close()  

@app.route('/Requisiciones')
def requisiciones():
    db = db_sql_server.get_connection()
    if db is not None:
        cursor = db.cursor()
        cursor.execute("SELECT ID, NOMBRE_PROYECTO FROM PROYECTOS")
        proyectos = cursor.fetchall()
    else:
        proyectos = []
    return render_template('requisiciones.html', proyectos=proyectos)

#Rutas Nomina
@app.route('/nominas', methods=['GET'])
def obtener_nominas():
    """
    Ruta para obtener todas las nóminas registradas.
    """
    try:
        # Obtener todas las nóminas usando el modelo
        resultado = ModelNomina.get_all_nominas(db)
        return jsonify([nomina.__dict__ for nomina in resultado]), 200
    except Exception as e:
        print(f"Error interno: {str(e)}")
        return jsonify({"error": f"Error interno del servidor: {str(e)}"}), 500
        

@app.route('/nominas/<int:id>', methods=['GET'])
def obtener_nomina_por_id(id):
    """
    Ruta para obtener una nómina específica por su ID.
    """
    try:
        resultado = ModelNomina.get_nomina_by_id(db, id)
        if resultado:
            return jsonify(resultado.__dict__), 200
        return jsonify({"error": "Nómina no encontrada"}), 404
    except Exception as e:
        print(f"Error interno: {str(e)}")
        return jsonify({"error": f"Error interno del servidor: {str(e)}"}), 500

@app.route('/crear_actualizar_nominas', methods=['POST'])
def crear_actualizar_nominas():
    try:
        data = request.get_json()

        for nomina_data in data.get('nominas', []):
            nomina = Nomina(
                id_empleado=nomina_data['id_empleado'],
                fecha_inicio=nomina_data['fecha_inicio'],
                fecha_final=nomina_data['fecha_final'],
                extra=nomina_data.get('extra', 0),
                deuda=nomina_data.get('deuda', 0),
                total_nomina=nomina_data['total_nomina'],
                fecha_registro=nomina_data['fecha_registro']
            )

            # Verificar si ya existe una nómina para ese empleado y semana
            existing_nomina = ModelNomina.get_nomina_by_id(db, nomina.id_empleado)
            if existing_nomina:
                ModelNomina.update_nomina(db, nomina)
            else:
                ModelNomina.create_nomina(db, nomina)

        return jsonify({"message": "Nóminas procesadas correctamente."}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@app.route('/actualizar_nomina_empleado', methods=['POST'])
def actualizar_nomina_empleado():
    """
    Endpoint para actualizar o crear un registro de nómina basado en el ID del empleado y rango de fechas.
    """
    try:
        # Obtener los datos enviados en el cuerpo de la solicitud
        data = request.get_json()
        print("[DEBUG] Datos recibidos en el backend:", data)  # Debug: Imprimir los datos recibidos

        # Extraer datos y validarlos
        id_empleado = data.get("id_empleado")
        fecha_inicio = data.get("fecha_inicio")
        fecha_final = data.get("fecha_final")
        extra = data.get("extra", 0)
        deuda = data.get("deuda", 0)
        observaciones = data.get("observaciones", "").strip()

        # Log de los datos antes de validarlos
        print("[DEBUG] Datos extraídos:")
        print(f"  - id_empleado: {id_empleado} (tipo: {type(id_empleado)})")
        print(f"  - fecha_inicio: {fecha_inicio} (tipo: {type(fecha_inicio)})")
        print(f"  - fecha_final: {fecha_final} (tipo: {type(fecha_final)})")
        print(f"  - extra: {extra} (tipo: {type(extra)})")
        print(f"  - deuda: {deuda} (tipo: {type(deuda)})")
        print(f"  - observaciones: {observaciones} (tipo: {type(observaciones)})")

        # Validaciones básicas
        if not id_empleado or not isinstance(id_empleado, int):
            raise ValueError("El ID del empleado no es válido.")
        if not fecha_inicio or not fecha_final:
            raise ValueError("Las fechas de inicio y final son obligatorias.")
        if not isinstance(extra, (int, float)):
            raise ValueError("El valor de 'extra' debe ser numérico.")
        if not isinstance(deuda, (int, float)):
            raise ValueError("El valor de 'deuda' debe ser numérico.")

        # Log antes de enviar los datos al modelo
        print("[DEBUG] Datos validados, enviando al modelo:")
        print(f"  - id_empleado: {id_empleado}")
        print(f"  - fecha_inicio: {fecha_inicio}")
        print(f"  - fecha_final: {fecha_final}")
        print(f"  - extra: {extra}")
        print(f"  - deuda: {deuda}")
        print(f"  - observaciones: {observaciones}")

        # Lógica para crear o actualizar el registro de nómina
        resultado = ModelNomina.upsert_nomina(
            db,
            id_empleado=id_empleado,
            fecha_inicio=fecha_inicio,
            fecha_final=fecha_final,
            extra=extra,
            deuda=deuda,
            observaciones=observaciones,
        )

        # Log del resultado del modelo
        print("[DEBUG] Resultado del modelo:", resultado)

        # Responder con éxito
        action = resultado.get("action")
        id_nomina = resultado.get("id_nomina")
        message = f"Nómina {'actualizada' if action == 'updated' else 'creada'} correctamente."

        print("[DEBUG] Respuesta final del endpoint:", {
            "success": True,
            "message": message,
            "id_nomina": id_nomina,
            "action": action
        })

        return jsonify({
            "success": True,
            "message": message,
            "id_nomina": id_nomina,
            "action": action
        }), 200

    except ValueError as ve:
        print("[ERROR] Error de validación:", str(ve))  # Debug: Imprimir errores de validación
        return jsonify({"error": f"Datos inválidos: {str(ve)}"}), 400
    except Exception as e:
        print("[ERROR] Error inesperado:", str(e))  # Debug: Imprimir cualquier otro error inesperado
        return jsonify({"error": f"Error inesperado: {str(e)}"}), 500



@app.route('/nominas', methods=['POST'])
def crear_nomina():
    """
    Ruta para crear una nueva nómina.
    """
    try:
        data = request.json
        nueva_nomina = Nomina(
            id=None,
            id_empleado=data.get('id_empleado'),
            fecha_inicio=data.get('fecha_inicio'),
            fecha_final=data.get('fecha_final'),
            extra=data.get('extra'),
            deuda=data.get('deuda'),
            total_nomina=data.get('total_nomina'),
            fecha_registro=data.get('fecha_registro')
        )
        ModelNomina.create_nomina(db, nueva_nomina)
        return jsonify({"message": "Nómina creada exitosamente"}), 201
    except Exception as e:
        print(f"Error interno: {str(e)}")
        return jsonify({"error": f"Error interno del servidor: {str(e)}"}), 500

@app.route('/api/actualizar_nomina', methods=['POST'])
def actualizar_nomina():
    data = request.json
    id_empleado = data.get("id_empleado")
    extra = data.get("extra", 0)
    deuda = data.get("deuda", 0)
    pago_neto_semanal = data.get("pago_neto_semanal", 0)

    if not id_empleado:
        return jsonify({"error": "ID de empleado requerido"}), 400

    # Lógica para actualizar o crear el registro en la base de datos
    try:
        registro_nomina = Nomina.query.filter_by(id_empleado=id_empleado).first()
        if not registro_nomina:
            # Crear un nuevo registro si no existe
            registro_nomina = Nomina(id_empleado=id_empleado, extra=extra, deuda=deuda, pago_neto_semanal=pago_neto_semanal)
            db.session.add(registro_nomina)
        else:
            # Actualizar el registro existente
            registro_nomina.extra = extra
            registro_nomina.deuda = deuda
            registro_nomina.pago_neto_semanal = pago_neto_semanal

        db.session.commit()
        return jsonify({"message": "Registro de nómina actualizado correctamente."}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@app.route('/nominas/<int:id>', methods=['DELETE'])
def eliminar_nomina(id):
    """
    Ruta para eliminar una nómina por su ID.
    """
    try:
        ModelNomina.delete_nomina(db, id)
        return jsonify({"message": "Nómina eliminada exitosamente"}), 200
    except Exception as e:
        print(f"Error interno: {str(e)}")
        return jsonify({"error": f"Error interno del servidor: {str(e)}"}), 500

#FinRutasNomina  

# Portal de asistencia
asistencia_model = AsistenciaModel(db)

@app.route('/asistencia', methods=['GET'])
def asistencia():
    return render_template('asistencia.html')

# API PARA AUTOCOMPLETAR EMPLEADOS
@app.route('/api/buscar_empleado/', methods=['GET'])
def buscar_empleado():
    try:
        # Obtener el parámetro de búsqueda desde la URL
        query = request.args.get('query', '').strip()
        
        # Reutilizar el método filter_empleados del modelo
        empleados = ModelEmpleado.filter_empleados(
            db=db,
            nombre=query,
            apellido=None,  # No buscamos por apellido directamente
            tipo_empleado=None,
            estado='0'  # Buscar solo empleados no bloqueados
        )
        
        # Formatear los datos para la respuesta JSON
        empleados_filtrados = [
            {
                'id': e.id,
                'nombre': e.nombre,
                'apellido': e.apellido,
                'puesto': e.puesto
            }
            for e in empleados
        ]
        
        # Retornar los datos filtrados
        return jsonify(empleados_filtrados), 200
    
    except Exception as e:
        # Manejar errores y devolver un mensaje de error
        return jsonify({'error': str(e)}), 500


        
@app.route('/api/buscar_empleado_nombre_apellido/', methods=['GET'])
def buscar_empleado_nombre_apellido():
    try:
        # Obtener el parámetro de búsqueda desde la URL
        query = request.args.get('query', '').strip()
        
        if not query:
            return jsonify({'error': 'El término de búsqueda es requerido'}), 400

        # Dividir el término de búsqueda en posibles nombre y apellido
        query_parts = query.split()
        nombre = query_parts[0] if len(query_parts) > 0 else None
        apellido = " ".join(query_parts[1:]) if len(query_parts) > 1 else None

        # Reutilizar el método filter_empleados del modelo
        empleados = ModelEmpleado.filter_empleados(
            db=db,
            nombre=nombre,
            apellido=apellido,
            tipo_empleado=None,
            estado='0'  # Buscar solo empleados no bloqueados
        )
        
        # Formatear los datos para la respuesta JSON
        empleados_filtrados = [
            {
                'id': e.id,
                'nombre': e.nombre,
                'apellido': e.apellido,
                'puesto': e.puesto,
                'nomina': e.nomina
            }
            for e in empleados
        ]
        
        if empleados_filtrados:
            return jsonify(empleados_filtrados), 200
        else:
            return jsonify({'error': 'No se encontraron empleados con el nombre y apellido especificados'}), 404

    except Exception as e:
        # Manejar errores y devolver un mensaje de error
        return jsonify({'error': str(e)}), 500


@app.route('/registrar_asistencia_empleado', methods=['POST'])
def registrar_asistencia_empleado():
    """
    Ruta para registrar la asistencia de un empleado.
    """
    try:
        data = request.get_json()

        # Validaciones básicas
        if not data:
            return jsonify({"error": "El cuerpo de la solicitud está vacío o no es JSON válido."}), 400

        campos_requeridos = ['id_empleado', 'id_proyecto', 'dia', 'hora_entrada', 'latitud', 'longitud', 'foto_base64']
        for campo in campos_requeridos:
            if campo not in data:
                return jsonify({"error": f"El campo '{campo}' es obligatorio."}), 400

        # Validar formato de fecha y hora
        try:
            data['dia'] = datetime.strptime(data['dia'], '%Y-%m-%d').date()
            data['hora_entrada'] = datetime.strptime(data['hora_entrada'], '%H:%M:%S').time()
        except ValueError:
            return jsonify({"error": "El formato de 'dia' debe ser 'YYYY-MM-DD' y el de 'hora_entrada' debe ser 'HH:MM:SS'."}), 400

        # Registrar asistencia utilizando el modelo
        resultado = asistencia_model.registrar_asistencia(data)

        if "error" in resultado:
            return jsonify(resultado), 500
        return jsonify(resultado), 200
    except Exception as e:
        print(f"Error interno: {str(e)}")
        return jsonify({"error": f"Error interno del servidor: {str(e)}"}), 500


@app.route('/registrar_hora_salida', methods=['POST'])
def registrar_hora_salida():
    """
    Ruta para registrar la hora de salida de un empleado.
    """
    try:
        data = request.get_json()

        if not data or 'id_empleado' not in data:
            return jsonify({"error": "El campo 'id_empleado' es obligatorio."}), 400

        # Actualizar hora de salida utilizando el modelo
        resultado = asistencia_model.actualizar_hora_salida(data['id_empleado'])

        if "error" in resultado:
            return jsonify(resultado), 404 if "No se encontró" in resultado["error"] else 500
        return jsonify(resultado), 200
    except Exception as e:
        print(f"Error interno: {str(e)}")
        return jsonify({"error": f"Error interno del servidor: {str(e)}"}), 500


@app.route('/verificar_asistencia', methods=['GET'])
def verificar_asistencia():
    """
    Ruta para verificar si un empleado ya registró asistencia hoy.
    """
    try:
        id_empleado = request.args.get('id_empleado')

        if not id_empleado:
            return jsonify({"error": "El campo 'id_empleado' es obligatorio."}), 400

        # Verificar asistencia utilizando el modelo
        resultado = asistencia_model.verificar_asistencia(id_empleado)

        if "error" in resultado:
            return jsonify(resultado), 500
        return jsonify(resultado), 200
    except Exception as e:
        print(f"Error interno: {str(e)}")
        return jsonify({"error": f"Error interno del servidor: {str(e)}"}), 500

import logging

# Cofigurar el logger (puedes configurarlo según tus necesidades)
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

@app.route('/api/empleados_por_categoria/<string:categoria>/<int:id_proyecto>', methods=['GET'])
def obtener_empleados_por_categoria(categoria, id_proyecto):
    if db is None:
        print("Error: Database connection failed")
        return jsonify({"error": "Database connection failed"}), 500

    try:
        # Imprimir los parámetros de entrada para verificar que están llegando correctamente
        print(f"Recibiendo solicitud con categoria: {categoria} e id_proyecto: {id_proyecto}")

        # Llamar a la función para obtener empleados
        empleados = ModelEmpleado.get_empleados_por_categoria(db, categoria, id_proyecto)

        # Verificar que los empleados se han recuperado correctamente
        print(f"Empleados recuperados: {len(empleados)} empleados encontrados.")

        # Procesar los empleados y convertirlos en formato JSON
        empleados_json = []
        for emp in empleados:
            print(f"Procesando empleado: {emp.id}, {emp.nombre} {emp.apellido}")

            # Aquí agregamos el campo `monedero` junto con los otros campos
            empleado_data = {
                "id": emp.id, 
                "nombre": emp.nombre, 
                "apellido": emp.apellido,
                "sueldo_imss": emp.sueldo_imss,
                "base_sueldo": emp.nomina,  # Asegúrate de que 'nomina' esté siendo accedido correctamente
                "monedero": emp.monedero    # Agregamos el campo monedero aquí
            }

            # Imprimir el JSON generado para cada empleado
            print(f"Datos JSON generado: {empleado_data}")
            empleados_json.append(empleado_data)

        # Verificar que los datos JSON se están generando correctamente
        print(f"Datos JSON completos: {empleados_json}")

        return jsonify(empleados_json), 200

    except Exception as e:
        # Imprimir el error para entender qué salió mal
        print(f"Error en la ruta /api/empleados_por_categoria: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/empleado_by_id/<int:id_empleado>', methods=['GET'])
def obtener_empleado_por_id(id_empleado):
    if db is None:
        print("Error: Database connection failed")
        return jsonify({"error": "Database connection failed"}), 500

    try:
        # Llamamos al método para obtener el empleado por su ID
        empleado = ModelEmpleado.get_empleado_by_id(db, id_empleado)

        # Si no se encuentra el empleado, devolver un error
        if not empleado:
            return jsonify({"error": "Empleado no encontrado"}), 404

        # Convertir los datos del empleado en formato JSON
        empleado_data = {
            "id": empleado.id,
            "nombre": empleado.nombre,
            "apellido": empleado.apellido,
            "sueldo_imss": empleado.sueldo_imss,
            "monedero": empleado.monedero,
            "base_sueldo": empleado.nomina,
            # Aquí puedes agregar más campos si es necesario
        }

        # Retornar los datos del empleado en formato JSON
        return jsonify(empleado_data), 200

    except Exception as e:
        print(f"Error en la ruta /api/empleado_by_id: {str(e)}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/datos_modal_nomina', methods=['GET'])
def get_datos_modal_nomina():
    """Ruta para obtener proyectos y empleados no bloqueados"""
    try:
        # Obtener los proyectos no bloqueados
        proyectos = ModelProyectoObra.get_proyectos_not_block(db)
        # Obtener todos los empleados
        empleados = ModelEmpleado.get_all_empleados(db)

        
        # Preparar los datos para el JSON
        proyectos_data = [
            {
                "id": proyecto.id,
                "nombre_proyecto": proyecto.nombre_proyecto,
                "centro_comercial": proyecto.centro_comercial
            } for proyecto in proyectos
        ]

        empleados_data = [
            {
                "id": empleado.id,
                "nombre_completo": f"{empleado.nombre} {empleado.apellido}",
                "puesto": empleado.puesto
            } for empleado in empleados
        ]

        # Devolver los datos en formato JSON
        return jsonify({
            "proyectos": proyectos_data,
            "empleados": empleados_data
        }), 200
    except Exception as ex:
        # Registrar el error en el log
        logging.error(f"Error al obtener los datos del modal: {str(ex)}")
        return jsonify({"error": str(ex)}), 500

@app.route('/api/buscar_proyectos_asignados/', methods=['GET'])
def buscar_proyectos_asignados():
    """
    Ruta para obtener los proyectos asignados a un empleado específico.
    """
    try:
        # Log: Inicio del endpoint
        app.logger.info("Iniciando búsqueda de proyectos asignados...")

        # Obtener parámetros de la solicitud
        id_empleado = request.args.get('id_empleado')
        query = request.args.get('query', '')

        # Validación del parámetro id_empleado
        if not id_empleado:
            app.logger.error("El ID del empleado no fue proporcionado.")
            return jsonify({'error': 'El ID del empleado es requerido'}), 400

        # Log: Parámetros recibidos
        app.logger.info(f"ID del empleado: {id_empleado}, Query: {query}")

        # Intentar obtener proyectos asignados
        proyectos = ModelAsignaciones.get_proyectos_asignados(db, int(id_empleado))
        app.logger.info(f"Proyectos encontrados: {len(proyectos)}")

        proyectos_filtrados = []

        # Filtrar proyectos por query
        for p in proyectos:
            if query.lower() in p.nombre_proyecto.lower():
                app.logger.info(f"Proyecto coincidente: {p.nombre_proyecto}")

                # Calcular semanas entre fecha de inicio y fin
                fecha_inicio = datetime.strptime(str(p.fecha_inicio), '%Y-%m-%d')
                fecha_fin = datetime.strptime(str(p.fecha_fin), '%Y-%m-%d')
                diferencia = fecha_fin - fecha_inicio
                semanas = round(diferencia.days / 7)

                # Construir dirección del proyecto
                direccion = f"{p.calle} {p.numero_exterior}"
                if p.numero_interior:
                    direccion += f" Int. {p.numero_interior}"
                direccion += f", {p.colonia}, {p.municipio}, {p.estado}, {p.pais}"

                proyectos_filtrados.append({
                    'id': p.id,
                    'nombre_proyecto': p.nombre_proyecto,
                    'fecha_inicio': p.fecha_inicio,
                    'fecha_fin': p.fecha_fin,
                    'semanas': semanas,
                    'direccion': direccion
                })

        # Log: Proyectos filtrados
        app.logger.info(f"Proyectos filtrados: {len(proyectos_filtrados)}")

        # Devolver respuesta
        return jsonify(proyectos_filtrados), 200

    except Exception as e:
        # Log: Error
        app.logger.error(f"Error en buscar_proyectos_asignados: {str(e)}")
        return jsonify({'error': 'Ocurrió un error interno en el servidor'}), 500



@app.route('/obtener_asistencias', methods=['GET'])
def obtener_asistencias():
    """
    Ruta para obtener todas las asistencias registradas.
    """
    try:
        # Obtener asistencias utilizando el modelo
        resultado = asistencia_model.obtener_asistencias()

        if "error" in resultado:
            return jsonify(resultado), 500
        return jsonify(resultado), 200
    except Exception as e:
        print(f"Error interno: {str(e)}")
        return jsonify({"error": f"Error interno del servidor: {str(e)}"}), 500
    
    
    
@app.route('/api/asistencias_general', methods=['GET'])
def obtener_asistencias_general():
    try:
        # Obtener parámetros de la solicitud
        fecha_inicio_str = request.args.get('fecha_inicio')
        fecha_fin_str = request.args.get('fecha_fin')
        id_proyecto = request.args.get('id_proyecto')  # Filtro opcional por proyecto

        # Manejar fechas
        if fecha_inicio_str and fecha_fin_str:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
        else:
            hoy = datetime.now().date()
            # Calcular rango de lunes a domingo basado en la fecha actual
            fecha_inicio = hoy - timedelta(days=hoy.weekday())  # Lunes
            fecha_fin = fecha_inicio + timedelta(days=6)  # Domingo

        # Validar id_proyecto como entero
        if id_proyecto:
            try:
                id_proyecto = int(id_proyecto)
            except ValueError:
                return jsonify({"error": "El parámetro id_proyecto debe ser un número entero."}), 400

        # Obtener asistencias del modelo
        asistencias = asistencia_model.obtener_asistencias_por_rango(fecha_inicio, fecha_fin, id_proyecto)

        # Validar que sea una lista y manejar errores
        if not isinstance(asistencias, list):
            print(f"Error en modelo: {asistencias.get('error', 'Error desconocido')}")
            return jsonify({"error": asistencias.get("error", "Error interno del servidor.")}), 500

        # Ajustar datos faltantes en asistencias
        for asistencia in asistencias:
            asistencia.setdefault("nombre_proyecto", "No asignado")  # Valor predeterminado

        return jsonify(asistencias), 200

    except Exception as e:
        print(f"Error en la ruta: {e}")
        return jsonify({"error": "Error interno del servidor."}), 500
    


@app.route('/api/asistencias_general2', methods=['GET'])
def obtener_asistencias_general2():
    try:
        # Obtener parámetros de la solicitud
        fecha_inicio_str = request.args.get('fecha_inicio')
        fecha_fin_str = request.args.get('fecha_fin')
        id_proyecto = request.args.get('id_proyecto')  # Filtro opcional por proyecto

        # Manejar fechas
        if fecha_inicio_str and fecha_fin_str:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
        else:
            hoy = datetime.now().date()
            fecha_inicio = hoy - timedelta(days=hoy.weekday())  # Lunes
            fecha_fin = fecha_inicio + timedelta(days=6)  # Domingo

        # Validar id_proyecto como entero
        if id_proyecto:
            try:
                id_proyecto = int(id_proyecto)
            except ValueError:
                return jsonify({"error": "El parámetro id_proyecto debe ser un número entero."}), 400

        # Obtener asistencias del modelo
        asistencias = asistencia_model.obtener_asistencias_por_rango2(fecha_inicio, fecha_fin, id_proyecto)

        # Validar que sea una lista y manejar errores
        if not isinstance(asistencias, list):
            print(f"⚠️ Error en modelo: {asistencias.get('error', 'Error desconocido')}")
            return jsonify({"error": asistencias.get("error", "Error interno del servidor.")}), 500

        # 🔹 **Calcular las horas extra por empleado**
        total_horas_extra_global = 0  # Para acumular el total de todas las horas extra

        for asistencia in asistencias:
            total_horas_extra = 0  # Reiniciar el contador por empleado

            for dia in ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]:
                if dia in asistencia:
                    hora_salida = asistencia[dia].get("hora_salida", None)
                    horario_salida = asistencia.get("horario_personalizado_salida") or asistencia.get("hora_salida_proyecto")

                    # 🕒 **Cálculo de horas extra solo si hay salida registrada**
                    if hora_salida and horario_salida:
                        try:
                            hora_salida_asistencia = datetime.strptime(hora_salida, "%H:%M")
                            hora_salida_horario = datetime.strptime(horario_salida, "%H:%M")

                            if hora_salida_asistencia > hora_salida_horario:
                                diferencia = hora_salida_asistencia - hora_salida_horario
                                horas_extra = diferencia.total_seconds() / 3600  # Convertir a horas
                                horas_extra = round(horas_extra, 2)  # Redondear a 2 decimales
                                total_horas_extra += horas_extra  # Sumar al total del empleado
                        except Exception as e:
                            print(f"⚠️ Error calculando horas extra para {asistencia['nombre_empleado']} en {dia}: {e}")

            # 🔥 **Añadir el total de horas extra al objeto del empleado**
            asistencia["total_horas_extra"] = total_horas_extra
            total_horas_extra_global += total_horas_extra  # Acumular en el total general

        # 🔥 **Imprimir solo el total global de horas extra**
        print(f"\n🕒 Total de horas extra trabajadas en el rango: {total_horas_extra_global} horas\n")

        return jsonify(asistencias), 200

    except Exception as e:
        print(f"❌ Error en la ruta: {e}")
        return jsonify({"error": "Error interno del servidor."}), 500


@app.route('/api/modificar_horas_extra', methods=['POST'])
def modificar_horas_extra():
    """
    Ruta para modificar las horas extras de una asistencia específica.
    """
    try:
        print("Iniciando la ruta modificar_horas_extra")
        data = request.json
        print(f"Datos recibidos: {data}")

        id_empleado = data.get('id_empleado')
        fecha = data.get('fecha')
        horas_extra = data.get('horas_extra')

        # Validar datos
        if not all([id_empleado, fecha, horas_extra is not None]):
            return jsonify({"error": "Faltan datos obligatorios (id_empleado, fecha, horas_extra)."}), 400

        # Validar tipo de datos
        try:
            horas_extra = float(horas_extra)
        except ValueError:
            return jsonify({"error": "El campo horas_extra debe ser un número válido."}), 400

        # Llamar a la función del modelo
        resultado = asistencia_model.actualizar_horas_extra(id_empleado, fecha, horas_extra)

        # Manejar respuesta del modelo
        if "error" in resultado:
            return jsonify(resultado), 400

        return jsonify(resultado), 200

    except Exception as e:
        print(f"Error en la ruta modificar_horas_extra: {str(e)}")
        return jsonify({"error": f"Error interno del servidor: {str(e)}"}), 500
    
@app.route('/asignar_empleado', methods=['POST'])
def asignar_empleado():
    """
    Ruta para asignar un empleado a un proyecto con sueldos, monedero y horarios opcionales.
    """
    data = request.json

    # Validar que se envíen los datos requeridos
    id_empleado = data.get('id_empleado')
    id_proyecto = data.get('id_proyecto')
    sueldo_base = data.get('sueldo_base')
    sueldo_imss = data.get('sueldo_imss')
    monedero = data.get('monedero')
    id_detalle_bauart = data.get('id_detalle_bauart')  # Para actualización de Detalle Bauart
    hora_entrada_p = data.get('hora_entrada_p')  # Campo opcional
    hora_salida_p = data.get('hora_salida_p')  # Campo opcional

    if not id_empleado or not id_proyecto or sueldo_base is None or sueldo_imss is None or monedero is None:
        return jsonify({'error': 'Faltan datos requeridos: id_empleado, id_proyecto, sueldo_base, sueldo_imss, monedero'}), 400

    try:
        # Validar existencia del empleado
        empleado = ModelEmpleado.get_empleado_by_id(db, id_empleado)
        if not empleado:
            return jsonify({'error': f'El empleado con ID {id_empleado} no existe.'}), 404

        # Validar existencia del proyecto
        proyecto = ModelProyectoObra.get_proyecto_by_id(db, id_proyecto)
        if not proyecto:
            return jsonify({'error': f'El proyecto con ID {id_proyecto} no existe.'}), 404

        # Siempre usar la fecha actual para la asignación
        fecha_asignacion = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Crear la asignación con los nuevos campos
        nueva_asignacion = Asignacion(
            id_empleado=id_empleado,
            id_proyecto=id_proyecto,
            sueldo_base=sueldo_base,
            sueldo_imss=sueldo_imss,
            monedero=monedero,
            fecha_asignacion=fecha_asignacion,  # 🔥 Siempre usamos la fecha actual aquí
            fecha_fin=data.get('fecha_fin', None),  # Fecha opcional
            id_detalle=id_detalle_bauart,  # Se añadió este campo
            hora_entrada_p=hora_entrada_p,  # Nuevo campo opcional
            hora_salida_p=hora_salida_p  # Nuevo campo opcional
        )

        # Guardar la asignación en la base de datos
        ModelAsignaciones.add_asignacion(db.cursor(), nueva_asignacion)
        db.commit()  # Confirmar la transacción

        # Si hay un id_detalle_bauart, actualizarlo con el ID de la asignación en lugar del ID del empleado
        if id_detalle_bauart:
            cursor = db.cursor()
            ModelAsignaciones.update_id_asignacion_detalle_bauart(cursor, id_detalle_bauart, nueva_asignacion.id_empleado)
            db.commit()  # Confirmar la transacción

        # Crear manualmente el diccionario de respuesta con los nuevos campos
        asignacion_data = {
            'id_empleado': nueva_asignacion.id_empleado,
            'id_proyecto': nueva_asignacion.id_proyecto,
            'sueldo_base': nueva_asignacion.sueldo_base,
            'sueldo_imss': nueva_asignacion.sueldo_imss,
            'monedero': nueva_asignacion.monedero,
            'fecha_asignacion': fecha_asignacion, 
            'fecha_fin': nueva_asignacion.fecha_fin,
            'id_detalle': nueva_asignacion.id_detalle,
            'hora_entrada_p': nueva_asignacion.hora_entrada_p,  # Nuevo campo en respuesta
            'hora_salida_p': nueva_asignacion.hora_salida_p  # Nuevo campo en respuesta
        }

        return jsonify({'message': 'Asignación creada exitosamente.', 'asignacion': asignacion_data}), 201

    except Exception as ex:
        db.rollback()  # Revertir cambios en caso de error
        return jsonify({'error': f'Error al asignar empleado: {str(ex)}'}), 500


@app.route('/desasignar_empleado', methods=['POST'])
def desasignar_empleado():
    """
    Ruta para desasignar un empleado de un proyecto, actualizando la fecha de fin en la asignación 
    y desasignando el detalle de Bauart.
    """
    data = request.json

    # Validar que se envíen los datos requeridos
    id_empleado = data.get('id_empleado')
    id_proyecto = data.get('id_proyecto')

    if not id_empleado or not id_proyecto:
        return jsonify({'error': 'Faltan datos requeridos: id_empleado, id_proyecto'}), 400

    try:
        # Validar existencia del empleado
        empleado = ModelEmpleado.get_empleado_by_id(db, id_empleado)
        if not empleado:
            return jsonify({'error': f'El empleado con ID {id_empleado} no existe.'}), 404

        # Validar existencia del proyecto
        proyecto = ModelProyectoObra.get_proyecto_by_id(db, id_proyecto)
        if not proyecto:
            return jsonify({'error': f'El proyecto con ID {id_proyecto} no existe.'}), 404

        # Desasignar el empleado
        cursor = db.cursor()
        if ModelAsignaciones.desasignar_empleado(cursor, id_empleado, id_proyecto):
            db.commit()  # Confirmar la transacción

            return jsonify({'message': 'Empleado desasignado exitosamente.'}), 200
        else:
            return jsonify({'error': 'Error al desasignar el empleado.'}), 500

    except Exception as ex:
        db.rollback()  # Revertir cambios en caso de error
        return jsonify({'error': f'Error al desasignar empleado: {str(ex)}'}), 500

@app.route('/detalles_proyecto/<int:id_proyecto>', methods=['GET'])
def obtener_detalles_proyecto(id_proyecto):

    try:
        with db.cursor() as cursor:

            detalles = ModelAsignaciones.get_project_details(cursor, id_proyecto)


        if not detalles:

            return jsonify({'error': f'No se encontraron detalles para el proyecto {id_proyecto}'}), 404

        return jsonify({'id_proyecto': id_proyecto, 'detalles': detalles}), 200

    except Exception as ex:
        import traceback
        print(f"❌ Error al obtener los detalles del proyecto {id_proyecto}:\n{traceback.format_exc()}")
        return jsonify({'error': f'Error al obtener los detalles del proyecto: {str(ex)}'}), 500



@app.route('/sabanaasistencias2', methods=['GET'])
def sabanaasistencias2():
    # Obtener los parámetros de la URL
    project_id = request.args.get('project_id', None)  # ID del proyecto
    start_date = request.args.get('start_date', None)  # Fecha de inicio de la semana
    end_date = request.args.get('end_date', None)  # Fecha de fin de la semana

        # Aquí pasamos project_id, start_date y end_date a la plantilla
    return render_template('sabanaasistencias_2.html', project_id=project_id, start_date=start_date, end_date=end_date)

@app.route('/sabanaasistencias3', methods=['GET'])
def sabanaasistencia3():
    return render_template('pago.html')


@app.route('/sabanaasistencias4', methods=['GET'])
def sabanaasistencia4():
    return render_template('dispersionmonedero.html')

@app.route('/sabanaasistencias5', methods=['GET'])
def sabanaasistencia5():
    return render_template('dispersionimss.html')

@app.route('/exportar_excel', methods=['POST'])
def exportar_excel():
    """
    Ruta para generar un archivo Excel con los datos enviados desde el frontend,
    ignorando la columna 'ACCIONES' y agregando el rango de fechas bajo el título del proyecto,
    sin afectar la tabla y agregando logo y datos adicionales sin sobreponer.
    """
    try:
        # Obtener los datos JSON enviados
        data = request.get_json()

        headers = data.get('headers')
        rows = data.get('rows')

        # Eliminar la columna 'ACCIONES' (suponiendo que la columna está en la última posición)
        if "ACCIONES" in headers:
            acciones_index = headers.index("ACCIONES")
            headers.pop(acciones_index)  # Eliminar 'ACCIONES' de los encabezados
            for row in rows:
                row.pop(acciones_index)  # Eliminar la columna 'ACCIONES' de cada fila

        # Verificar si los datos están vacíos
        if not headers or not rows:
            return {"error": "No hay datos para exportar"}, 400

        # Crear un DataFrame con los datos limpios
        df = pd.DataFrame(rows, columns=headers)

        # Crear un archivo Excel en memoria
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Asistencias")

        # Cargar el archivo Excel para aplicar los estilos
        output.seek(0)
        wb = load_workbook(output)
        sheet = wb["Asistencias"]

        # Insertar una fila en la parte superior para logo y otros datos
        sheet.insert_rows(1, 5)  # Inserta 5 filas para espacio para logo, nombre del proyecto y más

        # Insertar logo en la parte superior izquierda (A1)
        logo_path = 'static/img/Bauart-Logo-Pantalla-RGB.MED.JPG'  # Ruta al logo en tu carpeta estática
        img = Image(logo_path)
        img.width = 250  # Ajusta el tamaño horizontal del logo (más largo)
        img.height = 60  # Ajusta el alto para hacer el logo menos alto
        sheet.add_image(img, 'A1')  # Insertar la imagen en la celda A1

        # Mover el nombre del proyecto (ahora en D2)
        project_name = data.get('project_name', 'Proyecto sin nombre')  # Obtener el nombre del proyecto
        sheet['D2'] = f"{project_name}"
        sheet['D2'].font = Font(size=14, bold=True)
        sheet['D2'].alignment = Alignment(horizontal="center", vertical="center")

        # Mover la información de la empresa (ahora en D3)
        sheet['D3'] = "BA HOLDING, S.A. DE C.V."
        sheet['D3'].font = Font(size=12, italic=True)
        sheet['D3'].alignment = Alignment(horizontal="center", vertical="center")

        # Insertar el rango de fechas (ahora en D4)
        date_range = data.get('date_range', 'Rango no especificado')
        sheet['D4'] = f"Rango de Fechas: {date_range}"
        sheet['D4'].font = Font(size=12, italic=True)
        sheet['D4'].alignment = Alignment(horizontal="center", vertical="center")

        # Colocar Líder de Proyecto y Periodo en las celdas a la derecha (G1 y G2)
        leader_project = data.get('leader_project', 'Líder de Proyecto')
        period = data.get('period', 'Periodo no especificado')
        sheet['G1'] = f"LÍDER DE PROYECTO: {leader_project}"
        sheet['G1'].font = Font(size=12, bold=True)
        sheet['G1'].alignment = Alignment(horizontal="center", vertical="center")

        sheet['G2'] = f"PERIODO: {period}"
        sheet['G2'].font = Font(size=12, bold=True)
        sheet['G2'].alignment = Alignment(horizontal="center", vertical="center")

        # Aplicar color azul cielo a los encabezados y bordes fuertes
        blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        bold_font = Font(bold=True)
        black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        white_font = Font(bold=True, color="FFFFFF")
        alignment = Alignment(horizontal="center", vertical="center")

        # Añadir bordes más fuertes para los encabezados
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        border_strong = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

        # Aplicar formato de encabezado
        for cell in sheet[6]:  # Los encabezados están ahora en la fila 6
            if cell.value:  # Solo aplicar si hay un valor
                cell.fill = blue_fill
                cell.font = bold_font
                cell.border = border_strong  # Bordes más fuertes para los encabezados
                cell.alignment = alignment

        # Añadir bordes más claros a las celdas normales (sin encabezado)
        for row in sheet.iter_rows(min_row=7):  # Empezamos desde la fila 7 (datos)
            for cell in row:
                if cell.value:  # Solo aplicar bordes a celdas con contenido
                    cell.border = border_thin  # Bordes delgados para las celdas normales

        # Ajustar el ancho de las columnas automáticamente con un poco de espacio adicional
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Obtener la letra de la columna
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 5)
            sheet.column_dimensions[column].width = adjusted_width

        # Ajustar el texto en todas las celdas a centrado
        for row in sheet.iter_rows(min_row=7):  # Empezamos desde la fila 7 (datos)
            for cell in row:
                cell.alignment = alignment

        # Crear nombre de archivo con el nombre del proyecto y rango de fechas
        file_name = f"{project_name}_{date_range.replace(' ', '_').replace(':', '').replace('-', '_')}.xlsx"

        # Guardar el archivo modificado en memoria
        output.seek(0)
        wb.save(output)

        output.seek(0)

        # Enviar el archivo Excel como respuesta
        return send_file(output, as_attachment=True, download_name=file_name, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        # Capturar cualquier excepción y mostrar el error
        return {"error": f"Error al generar el archivo Excel: {str(e)}"}, 500

#RUTAS PDF 

@app.route('/get_partidas/<int:requisicion_id>', methods=['GET'])
def get_partidas(requisicion_id):
    db = db_sql_server.get_connection()
    if db is None:
        return jsonify({"error": "Database connection failed"}), 500

    try:
        with db.cursor() as cursor:
            # Consulta ajustada con JOIN a CATALOGO_MATERIALES_FAMILIAS
            query = """
                SELECT 
                    pr.ID, 
                    cmf.MATERIAL AS nombre_concepto,  -- Obtener el nombre del concepto
                    cmf.UNIDAD_MEDIDA AS unidad_medida,  -- Unidad de medida
                    pr.CANTIDAD, 
                    pr.FECHA_CREACION, 
                    pr.DETALLES
                FROM 
                    PARTIDAS_REQUISICION pr
                LEFT JOIN 
                    CATALOGO_MATERIALES_FAMILIAS cmf ON pr.DESCRIPCION = cmf.ID 
                WHERE 
                    pr.ID_REQUISICION = ?
            """
            cursor.execute(query, (requisicion_id,))
            rows = cursor.fetchall()

            partidas = []
            for row in rows:
                # Formatear los datos
                fecha_creacion = row[4].strftime("%d-%m-%Y %H:%M") if row[4] else "N/A"
                partidas.append({
                    "id": row[0],
                    "descripcion": row[1],  # Nombre del concepto/material
                    "unidad": row[2],       # Unidad de medida
                    "cantidad": row[3],
                    "fecha_creacion": fecha_creacion,
                    "detalles": row[5]
                })
            return jsonify({"partidas": partidas})
    except Exception as e:
        print(f"Error fetching partition data: {e}")
        return jsonify({"error": "Error fetching partition data"}), 500
    finally:
        db.close()

        
## MANEJO DE ERRORES
def status_401(error):
    return render_template('error401.html')

def status_404(error):
    return render_template('error404.html')

app.register_error_handler(401,status_401)
app.register_error_handler(404,status_404)

if __name__ == '__main__':
    app.run(debug=True)
    #app.config.from_object(config['development'])
    
    
