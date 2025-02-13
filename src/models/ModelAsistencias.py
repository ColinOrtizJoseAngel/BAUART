from datetime import datetime, timedelta
from math import radians, sin, cos, sqrt, atan2
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from io import StringIO


class AsistenciaModel:
    def __init__(self, db_connection):
        self.db_connection = db_connection

    def _calcular_distancia(self, lat1, lon1, lat2, lon2):
        """
        Calcula la distancia entre dos puntos geográficos utilizando la fórmula haversine.
        """
        R = 6371  # Radio de la Tierra en kilómetros
        lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
        dlat = lat2 - lat1
        dlon = lon2 - lon1
        a = sin(dlat / 2) ** 2 + cos(lat1) * cos(lat2) * sin(dlon / 2) ** 2
        c = 2 * atan2(sqrt(a), sqrt(1 - a))
        return R * c

    def registrar_asistencia(self, asistencia):
        """
        Registra la asistencia del empleado si está a menos de 2 km del proyecto.
        """
        try:
            query_proyecto = """
                SELECT LATITUD, LONGITUD 
                FROM Proyectos 
                WHERE ID = ?
            """
            with self.db_connection.cursor() as cursor:
                cursor.execute(query_proyecto, (asistencia['id_proyecto'],))
                proyecto = cursor.fetchone()
                if not proyecto:
                    return {"error": "El proyecto seleccionado no existe."}

                lat_proyecto = float(proyecto.LATITUD)
                lon_proyecto = float(proyecto.LONGITUD)

            lat_empleado = float(asistencia['latitud'])
            lon_empleado = float(asistencia['longitud'])
            distancia = self._calcular_distancia(
                lat_proyecto, lon_proyecto, lat_empleado, lon_empleado
            )

            if distancia > 100000:
                return {"error": f"La distancia al proyecto es de {distancia:.100000f} km. Solo puedes registrar asistencia si estás a menos de 2 km."}

            query = """
                INSERT INTO Asistencia (ID_EMPLEADO, ID_PROYECTO, DIA, HORA_ENTRADA, LATITUD, LONGITUD, FOTO_BASE64)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """
            with self.db_connection.cursor() as cursor:
                cursor.execute(
                    query,
                    (
                        asistencia['id_empleado'],
                        asistencia['id_proyecto'],
                        asistencia['dia'],
                        asistencia['hora_entrada'],
                        asistencia['latitud'],
                        asistencia['longitud'],
                        asistencia['foto_base64'],
                    )
                )
            self.db_connection.commit()
            return {"mensaje": "Asistencia registrada exitosamente"}
        except Exception as e:
            print(f"Error al registrar asistencia: {str(e)}")
            return {"error": str(e)}

    def actualizar_hora_salida(self, id_empleado):
        """
        Actualiza la hora de salida del empleado.
        """
        try:
            hora_salida = datetime.now().strftime('%H:%M:%S')

            # Verificar si ya existe un registro para hoy
            query_verificar = """
                SELECT HORA_SALIDA
                FROM Asistencia
                WHERE ID_EMPLEADO = ? AND DIA = CAST(GETDATE() AS DATE)
            """
            with self.db_connection.cursor() as cursor:
                cursor.execute(query_verificar, (id_empleado,))
                row = cursor.fetchone()

                if not row:
                    return {"error": "No se encontró un registro pendiente de salida para este empleado hoy."}
                if row.HORA_SALIDA is not None:
                    return {"error": "La hora de salida ya ha sido registrada para este empleado hoy."}

            # Actualizar HORA_SALIDA si está pendiente
            query_actualizar = """
                UPDATE Asistencia
                SET HORA_SALIDA = ?
                WHERE ID_EMPLEADO = ? AND DIA = CAST(GETDATE() AS DATE) AND HORA_SALIDA IS NULL
            """
            with self.db_connection.cursor() as cursor:
                cursor.execute(query_actualizar, (hora_salida, id_empleado))
                if cursor.rowcount == 0:
                    return {"error": "No se encontró un registro pendiente de salida para este empleado hoy."}

            self.db_connection.commit()
            return {"mensaje": "Hora de salida registrada correctamente"}
        except Exception as e:
            print(f"Error al actualizar hora de salida: {str(e)}")
            return {"error": str(e)}

    def verificar_asistencia(self, id_empleado):
        """
        Verifica si el empleado ha registrado asistencia hoy.
        """
        try:
            query = """
                SELECT HORA_ENTRADA, HORA_SALIDA
                FROM Asistencia
                WHERE ID_EMPLEADO = ? AND DIA = CAST(GETDATE() AS DATE)
            """
            with self.db_connection.cursor() as cursor:
                cursor.execute(query, (id_empleado,))
                row = cursor.fetchone()
                if row:
                    return {
                        "asistencia_registrada": True,
                        "hora_entrada": str(row.HORA_ENTRADA),
                        "hora_salida": str(row.HORA_SALIDA) if row.HORA_SALIDA else None,
                    }
                else:
                    return {"asistencia_registrada": False}
        except Exception as e:
            print(f"Error al verificar asistencia: {str(e)}")
            return {"error": str(e)}

    

    def obtener_asistencias_por_rango(self, fecha_inicio, fecha_fin, id_proyecto=None):
        try:
            query = """
                    SELECT 
                    e.ID AS empleado_id,
                    CONCAT(e.NOMBRE, ' ', e.APELLIDO) AS nombre_completo,
                    p.NOMBRE_PROYECTO AS nombre_proyecto,
                    p.FECHA_INICIO AS fecha_inicio_proyecto,
                    p.FECHA_FIN AS fecha_fin_proyecto,
                    e.CATEGORIA AS categoria,
                    e.TIPO_NOMINA AS tipo_nomina, 
                    asign.FECHA_ASIGNACION AS fecha_asignacion, 
                    asign.hora_entrada_p AS horario_personalizado_entrada,  
                    asign.hora_salida_p AS horario_personalizado_salida,  
                    a.DIA AS fecha,
                    a.HORA_ENTRADA AS hora_entrada,
                    a.HORA_SALIDA AS hora_salida,
                    a.HORAS_EXTRA AS horas_extra,
                    a.FOTO_BASE64 AS foto_base64,
                    a.LATITUD AS latitud,
                    a.LONGITUD AS longitud,
                    e.TOPE_HORAS_EXTRA AS tope_horas_extra,  
                    n.HORAS_EXTRA AS cantidad_horas_extra,  
                    n.EXTRA AS extra,  
                    n.DEUDA AS deuda,
                    asign.SUELDO_IMSS AS sueldo_imss,
                    asign.SUELDO_BASE AS nomina,
                    a.ID_Incidencia AS id_incidencia,
                    a.Es_Incidencia AS es_incidencia,
                    i.TipoIncidencia AS tipo_incidencia
                FROM Empleados e
                LEFT JOIN ASIGNACIONES asign ON e.ID = asign.ID_EMPLEADO
                LEFT JOIN Proyectos p ON asign.ID_PROYECTO = p.ID
                LEFT JOIN Asistencia a ON e.ID = a.ID_EMPLEADO AND a.DIA BETWEEN ? AND ?
                LEFT JOIN Nomina n ON e.ID = n.ID_Empleado  
                    AND n.Fecha_Inicio <= ?  
                    AND n.Fecha_Final >= ?  
                LEFT JOIN Incidencias i ON a.ID_Incidencia = i.IdIncidencia
            """

            params = [fecha_inicio, fecha_fin, fecha_inicio, fecha_fin]

            if id_proyecto:
                query += " WHERE asign.ID_PROYECTO = ?"
                params.append(id_proyecto)

            query += " ORDER BY e.NOMBRE, a.DIA"

            with self.db_connection.cursor() as cursor:
                cursor.execute(query, tuple(params))
                rows = cursor.fetchall()

                dias_semana = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
                asistencias = {}
                fecha_actual = datetime.now().date()

                for row in rows:
                    empleado_id = row.empleado_id
                    fecha_asignacion = row.fecha_asignacion

                    sueldo_base = row.nomina or 0.0

                    if row.tipo_nomina == 2:
                        sueldo_base = row.nomina
                    elif row.tipo_nomina == 1:
                        sueldo_base /= 2

                    if isinstance(fecha_asignacion, datetime):
                        fecha_asignacion = fecha_asignacion.date()
                    if row.fecha and isinstance(row.fecha, datetime):
                        row.fecha = row.fecha.date()

                    if empleado_id not in asistencias:
                        asistencias[empleado_id] = {
                            "id_empleado": empleado_id,
                            "nombre_empleado": row.nombre_completo,
                            "nombre_proyecto": row.nombre_proyecto or "No asignado",
                            "nomina": sueldo_base,
                            "categoria": row.categoria or "Sin categoría",
                            "tope_horas_extra": row.tope_horas_extra or 0,
                            "cantidad_horas_extra": row.cantidad_horas_extra or 0,
                            "extra": row.extra or 0,
                            "deuda": row.deuda or 0,
                            "horario_personalizado_entrada": row.horario_personalizado_entrada.strftime("%H:%M") if row.horario_personalizado_entrada else None,
                            "horario_personalizado_salida": row.horario_personalizado_salida.strftime("%H:%M") if row.horario_personalizado_salida else None
                        }

                        for i in range(7):
                            dia = fecha_inicio + timedelta(days=i)
                            if dia < fecha_asignacion:
                                estado = "N/A"
                            else:
                                estado = "N/A" if dia > fecha_actual else "F"

                            asistencias[empleado_id][dias_semana[i]] = {
                                "estado": estado,
                                "hora_entrada": None,
                                "hora_salida": None,
                                "horas_extra": None,
                                "foto_base64": None,
                                "latitud": None,
                                "longitud": None,
                                "tope_horas_extra": row.tope_horas_extra or 0,
                                "cantidad_horas_extra": row.cantidad_horas_extra or 0,
                                "extra": row.extra or 0,
                                "deuda": row.deuda or 0,
                                "id_incidencia": None,
                                "es_incidencia": None,
                                "tipo_incidencia": None
                            }

                    if row.fecha:
                        delta_dias = (row.fecha - fecha_inicio).days
                        if 0 <= delta_dias < 7:
                            dia_semana = dias_semana[delta_dias]
                            estado = "A" if row.hora_entrada and row.hora_salida else "F"

                            if row.es_incidencia == 1 and row.tipo_incidencia:
                                estado = row.tipo_incidencia.upper()

                            asistencias[empleado_id][dia_semana] = {
                                "estado": estado,
                                "hora_entrada": row.hora_entrada.strftime("%H:%M") if row.hora_entrada else None,
                                "hora_salida": row.hora_salida.strftime("%H:%M") if row.hora_salida else None,
                                "horas_extra": row.horas_extra,
                                "foto_base64": row.foto_base64,
                                "latitud": row.latitud,
                                "longitud": row.longitud,
                                "dia": row.fecha,
                                "tope_horas_extra": row.tope_horas_extra or 0,
                                "cantidad_horas_extra": row.cantidad_horas_extra or 0,
                                "extra": row.extra or 0,
                                "deuda": row.deuda or 0,
                                "id_incidencia": row.id_incidencia,
                                "es_incidencia": row.es_incidencia,
                                "tipo_incidencia": row.tipo_incidencia
                            }

            return list(asistencias.values())
        
        except Exception as e:
            print(f"Error al obtener asistencias por rango: {str(e)}")
            return {"error": str(e)}


    def obtener_asistencias_por_rango2(self, fecha_inicio, fecha_fin, id_proyecto=None):
        try:
            query = """
                SELECT 
                    e.ID AS empleado_id,
                    CONCAT(e.NOMBRE, ' ', e.APELLIDO) AS nombre_completo,
                    p.NOMBRE_PROYECTO AS nombre_proyecto,
                    p.HORA_ENTRADA AS hora_entrada_proyecto,
                    p.HORA_SALIDA AS hora_salida_proyecto,
                    e.TIPO_NOMINA AS tipo_nomina,
                    e.TOPE_HORAS_EXTRA AS tope_horas,
                    e.CATEGORIA AS categoria,
                    a.DIA AS fecha,
                    a.HORA_ENTRADA AS hora_entrada,
                    e.CLABE AS clabe,
                    a.HORA_SALIDA AS hora_salida,
                    a.HORAS_EXTRA AS horas_extra_asistencia,
                    a.FOTO_BASE64 AS foto_base64,
                    a.LATITUD AS latitud,
                    a.LONGITUD AS longitud,
                    a.ID_Incidencia AS id_incidencia,
                    a.Es_Incidencia AS es_incidencia,
                    i.TipoIncidencia AS tipo_incidencia,
                    n.Extra AS extra,
                    n.Deuda AS deuda,
                    n.Horas_Extra AS horas_extra_nomina,
                    n.Observaciones AS observaciones,
                    aas.MONEDERO AS monedero,
                    aas.SUELDO_IMSS AS sueldo_imss,
                    aas.SUELDO_BASE AS sueldo_base,
                    p.FECHA_INICIO AS fecha_inicio_proyecto,
                    p.FECHA_FIN AS fecha_fin_proyecto,
                    aas.FECHA_ASIGNACION as fecha_asignacion,
                    aas.hora_entrada_p as horario_personalizado_entrada,
                    aas.hora_salida_p as horario_personalizado_salida
                FROM Empleados e
                LEFT JOIN Asistencia a ON e.ID = a.ID_EMPLEADO AND a.DIA BETWEEN ? AND ?
                LEFT JOIN Proyectos p ON a.ID_PROYECTO = p.ID
                LEFT JOIN Incidencias i ON a.ID_Incidencia = i.IdIncidencia
                LEFT JOIN Nomina n ON e.ID = n.ID_Empleado 
                    AND n.Fecha_Inicio <= ?  
                    AND n.Fecha_Final >= ?  
                LEFT JOIN ASIGNACIONES aas ON e.ID = aas.ID_EMPLEADO
            """
            params = [fecha_inicio, fecha_fin, fecha_inicio, fecha_fin]

            if id_proyecto:
                query += " WHERE aas.ID_PROYECTO = ?"
                params.append(id_proyecto)

            query += " ORDER BY e.NOMBRE, a.DIA"

            with self.db_connection.cursor() as cursor:
                cursor.execute(query, tuple(params))
                rows = cursor.fetchall()

                dias_semana = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
                asistencias = {}

                for row in rows:
                    empleado_id = row.empleado_id
                    if empleado_id not in asistencias:
                        sueldo_imss = row.sueldo_imss or 0
                        monedero = row.monedero or 0
                        sueldo_base = row.sueldo_base or 0.0

                        if row.tipo_nomina == 2:
                            sueldo_base = row.sueldo_base
                        elif row.tipo_nomina == 1:
                            sueldo_base /= 2
                            sueldo_imss /= 2
                            monedero /= 2

                        extra = row.extra if row.extra is not None else 0.0
                        deuda = row.deuda if row.deuda is not None else 0.0
                        
                        monedero = float(monedero) + float(extra)
                        monedero = float(monedero) - float(deuda)
                        
                        asistencias[empleado_id] = {
                            "id_empleado": empleado_id,
                            "nombre_empleado": row.nombre_completo,
                            "nombre_proyecto": row.nombre_proyecto or "No asignado",
                            "categoria": row.categoria or "Sin categoría",
                            "tope_horas": row.tope_horas or 0,
                            "tipo_nomina": row.tipo_nomina or "No especificado",
                            "sueldo_imss": sueldo_imss,
                            "sueldo_base": sueldo_base,
                            "monedero": monedero,
                            "pago_monedero_neto": monedero,
                            "clabe": getattr(row, "clabe", "Sin CLABE"),
                            "extra": extra,
                            "deuda": deuda,
                            "observaciones": row.observaciones or "Sin observaciones",
                            "fecha_inicio_proyecto": row.fecha_inicio_proyecto,
                            "fecha_fin_proyecto": row.fecha_fin_proyecto,
                            "fecha_asignacion": row.fecha_asignacion,
                            "horario_personalizado_entrada": row.horario_personalizado_entrada.strftime("%H:%M") if row.horario_personalizado_entrada else None,
                            "horario_personalizado_salida": row.horario_personalizado_salida.strftime("%H:%M") if row.horario_personalizado_salida else None,
                            "hora_entrada_proyecto": row.hora_entrada_proyecto.strftime("%H:%M") if row.hora_entrada_proyecto else None,
                            "hora_salida_proyecto": row.hora_salida_proyecto.strftime("%H:%M") if row.hora_salida_proyecto else None,
                        }

                        for i in range(7):
                            dia = fecha_inicio + timedelta(days=i)
                            estado = "N/A" if dia > datetime.now().date() else "F"
                            asistencias[empleado_id][dias_semana[i]] = {
                                "estado": estado,
                                "hora_entrada": None,
                                "hora_salida": None,
                                "horas_extra": None,
                                "foto_base64": None,
                                "latitud": None,
                                "longitud": None,
                                "id_incidencia": None,
                                "es_incidencia": None,
                            }

                    if row.fecha:
                        delta_dias = (row.fecha - fecha_inicio).days
                        if 0 <= delta_dias < 7:
                            dia_semana = dias_semana[delta_dias]
                            estado = "A" if row.hora_entrada and row.hora_salida else "F"

                            if row.es_incidencia == 1 and row.tipo_incidencia:
                                estado = row.tipo_incidencia.upper()

                            horas_extra = row.horas_extra_asistencia or 0
                            extra = row.extra or 0
                            deuda = row.deuda or 0

                            asistencias[empleado_id][dia_semana] = {
                                "estado": estado,
                                "hora_entrada": row.hora_entrada.strftime("%H:%M") if row.hora_entrada else None,
                                "hora_salida": row.hora_salida.strftime("%H:%M") if row.hora_salida else None,
                                "horas_extra": horas_extra,
                                "foto_base64": row.foto_base64,
                                "latitud": row.latitud,
                                "longitud": row.longitud,
                                "dia": row.fecha,
                                "id_incidencia": row.id_incidencia,
                                "es_incidencia": row.es_incidencia,
                                "sueldo_real": row.sueldo_base
                            }

            return list(asistencias.values())

        except Exception as e:
            print(f"Error al obtener asistencias por rango: {str(e)}")
            return {"error": str(e)}



    def actualizar_horas_extra(self, id_empleado, fecha, horas_extra):
        """
        Actualiza las horas extras de una asistencia específica para un empleado en una fecha.

        Args:
            id_empleado (int): ID del empleado.
            fecha (date): Fecha de la asistencia.
            horas_extra (float): Cantidad de horas extras a registrar.

        Returns:
            dict: Mensaje de éxito o error.
        """
        try:
            with self.db_connection as conn:
                cursor = conn.cursor()

                # Verificar si existe el registro de asistencia para esa fecha
                query_verificar = """
                    SELECT ID FROM Asistencia WHERE ID_EMPLEADO = ? AND DIA = ?
                """
                cursor.execute(query_verificar, (id_empleado, fecha))
                asistencia = cursor.fetchone()

                if not asistencia:
                    return {"error": "No se encontró un registro de asistencia para esta fecha."}

                # Actualizar las horas extras
                query_actualizar = """
                    UPDATE Asistencia SET HORAS_EXTRA = ? WHERE ID_EMPLEADO = ? AND DIA = ?
                """
                cursor.execute(query_actualizar, (horas_extra, id_empleado, fecha))

                if cursor.rowcount == 0:
                    return {"error": "No se pudo actualizar las horas extras."}

                conn.commit()
                return {"mensaje": "Horas extras actualizadas correctamente"}

        except Exception as e:
            print(f"Error al actualizar horas extras: {str(e)}")
            return {"error": str(e)}

    def exportar_html_a_excel(self, html_file_path):
        """
        Lee una tabla HTML y la exporta a un archivo Excel con estilo.
        """
        try:
            # Leer la tabla desde el archivo HTML usando pandas
            df = pd.read_html(html_file_path)[0]  # Usamos la primera tabla

            # Crear un nuevo libro de trabajo de Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Asistencias"

            # Escribir los datos en la primera fila (encabezados)
            for col_num, column_title in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_num, value=column_title)
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Fondo amarillo
                cell.font = cell.font.copy(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Escribir los datos en las siguientes filas
            for row_num, row in enumerate(df.values, 2):
                for col_num, value in enumerate(row, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Fondo gris claro
                    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                         top=Side(style='thin'), bottom=Side(style='thin'))
                    cell.border = thin_border

            # Guardar el archivo Excel
            output_file = f"/tmp/asistencias_exportada.xlsx"
            wb.save(output_file)
            return {"mensaje": f"Archivo Excel generado correctamente: {output_file}"}

        except Exception as e:
            print(f"Error al exportar HTML a Excel: {str(e)}")
            return {"error": str(e)}